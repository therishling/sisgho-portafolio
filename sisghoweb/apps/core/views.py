from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect, HttpResponseBadRequest, JsonResponse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView, View, FormView
from django.views.generic.edit import ModelFormMixin
from django.urls import reverse, reverse_lazy
from apps.core import models as modelos
from apps.core import forms as formularios
from django.contrib.auth.mixins import UserPassesTestMixin, LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.db import connection
from datetime import date, datetime
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from itertools import chain
from django.db.models import Q
from django.views.decorators.http import require_http_methods
from django.template import Context
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from django.contrib import messages
from django.conf import settings
import os
from django.contrib.auth.forms import PasswordChangeForm
from django.core.mail import send_mail, EmailMultiAlternatives
import requests
import json
from django.template.loader import render_to_string


# PRINCIPALES

class Index(TemplateView):
    template_name = 'index.html'


class Dashboard(TemplateView):
    template_name = 'dashboard/index.html'

   

# Registrar Huespedes

class ReservarHabitacion(UserPassesTestMixin, SuccessMessageMixin, CreateView):
    model = modelos.Huesped
    form_class = formularios.HuespedForm
    template_name = 'dashboard/cliente/reservahabitacion.html'
    success_message = 'Huesped agregado.'
    success_url = reverse_lazy('reservar habitacion')

    def test_func(self):
        user = self.request.user
        if user.tipousuario.idtipousuario == 3:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')

    def get_context_data(self, **kwargs):

        habitaciones_disponible = len(modelos.Habitacion.objects.all().filter(estadohabitacion = 1))
        huespedes_sin_hab = len(modelos.Huesped.objects.all().filter(habitacion__isnull = True))
        hay_disponible = False
        if habitaciones_disponible > 0:
            hay_disponible = True
            if huespedes_sin_hab < habitaciones_disponible:
                hay_disponible = True
            else:
                hay_disponible = False
            



        context = super(ReservarHabitacion, self).get_context_data(**kwargs)
        context['huespedes'] = modelos.Huesped.objects.all().order_by('idhuesped')
        context['cliente'] = modelos.Cliente.objects.get(
            usuario=self.request.user.idusuario)
        context['time'] = date.today()
        print()
        context['hay_disponible'] = hay_disponible
        
        return context

    def form_valid(self, form, **kwargs):
        self.object = form.save(commit=False)

        if self.object.fechadesde < date.today():
            messages.error(self.request,'La "Fecha Desde", no puede ser menor a la fecha actual.')
            return HttpResponseRedirect(reverse('reservar habitacion'))
        
        if self.object.fechahasta < date.today():
            messages.error(self.request,'La "Fecha Hasta", no puede ser menor a la fecha actual.')
            return HttpResponseRedirect(reverse('reservar habitacion'))

           
        self.object = form.save()
        return super(ReservarHabitacion, self).form_valid(form)


class ModificarHuesped(UserPassesTestMixin, SuccessMessageMixin, UpdateView):

    model = modelos.Huesped
    form_class = formularios.HuespedForm
    template_name = 'dashboard/cliente/reservahabitacion.html'
    success_message = 'Huesped modificado.'
    success_url = reverse_lazy('reservar habitacion')
    context_object_name = 'huesped'

    def test_func(self):
        user = self.request.user
        if user.tipousuario.idtipousuario == 3:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')

    def form_valid(self, form, **kwargs):
        self.object = form.save(commit=False)

        if self.object.fechadesde < date.today():
            messages.error(self.request,'La "Fecha Desde", no puede ser menor a la fecha actual.')
            return HttpResponseRedirect(reverse('reservar habitacion'))
        
        if self.object.fechahasta < date.today():
            messages.error(self.request,'La "Fecha Hasta", no puede ser menor a la fecha actual.')
            return HttpResponseRedirect(reverse('reservar habitacion'))

           
        self.object = form.save()
        return super(ModificarHuesped, self).form_valid(form)


class CancelarReserva(UserPassesTestMixin, SuccessMessageMixin, DeleteView):
    model = modelos.Huesped
    template_name = 'dashboard/cliente/reservahabitacion.html'
    success_message = 'Reserva Cancelada'
    success_url = reverse_lazy('reservar habitacion')
    context_object_name = 'huesped'

    def test_func(self):
        user = self.request.user
        if user.tipousuario.idtipousuario == 3:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')


class SeleccionarHabitacion(UserPassesTestMixin, SuccessMessageMixin, CreateView):
    model = modelos.DetalleReserva
    form_class = formularios.DetalleReservaForm
    template_name = 'dashboard/empleado/asignarhabi.html'
    success_message = 'Habitacion asignada.'
    success_url = reverse_lazy('asignar habitacion')

    def test_func(self):
        user = self.request.user
        if user.tipousuario.idtipousuario == 2:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')

    def get_context_data(self, **kwargs):

        habitaciones_disponible = len(modelos.Habitacion.objects.all().filter(estadohabitacion = 1))
        hay_disponible = False
        if habitaciones_disponible > 0:
            hay_disponible = True
            
        
        context = super(SeleccionarHabitacion, self).get_context_data(**kwargs)
        context['huespedes'] = modelos.Huesped.objects.all().filter(fechahasta__gte = date.today()).order_by('idhuesped')
        context['habitaciones'] = modelos.Habitacion.objects.all().filter(
            estadohabitacion=1)
        context['hay_disponible'] = hay_disponible
        return context

    def post(self, request, *args, **kwargs):
        habitacion = ""
        huesped = ""
        if self.request.method == 'POST':
            huesped = modelos.Huesped.objects.get(
                idhuesped=self.request.POST.get('huesped'))
            habitacion = modelos.Habitacion.objects.get(
                idhabitacion=self.request.POST.get('habitacion'))
            huesped.habitacion = habitacion
            empleado = modelos.Empleado.objects.get(usuario = self.request.user.idusuario)
            huesped.empleado = empleado
       
            huesped.save()
            if self.request.POST.get('habitacion'):
                habitacion = modelos.Habitacion.objects.get(
                    idhabitacion=self.request.POST.get('habitacion'))
                estadohabitacion = modelos.Estadohabitacion.objects.get(
                    idestado=2)
                habitacion.estadohabitacion = estadohabitacion
                habitacion.save()

        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        huesped = modelos.Huesped.objects.get(
            idhuesped=self.request.POST.get('huesped'))

        habitacion = modelos.Habitacion.objects.get(
            idhabitacion=self.request.POST.get('habitacion'))
        dias = huesped.fechahasta - huesped.fechadesde
        total = dias.days * habitacion.precio
        self.object = form.save(commit=False)
        self.object.huesped = huesped
        self.object.dias = dias.days
        self.object.total = total
        self.object.habitacion = habitacion
        self.object = form.save()
        
        # Llamado a procedimiento almacenado
        with connection.cursor() as cursor:
            cursor.callproc('p_actualizar_habitaciones')
        return super(SeleccionarHabitacion, self).form_valid(form)


class AsignarHabitacion(UserPassesTestMixin, SuccessMessageMixin, UpdateView):
    model = modelos.Huesped
    form_class = formularios.AsHabitacionForm
    template_name = 'dashboard/empleado/asignarhabi.html'
    success_message = 'Habitacion asignada.'
    success_url = reverse_lazy('asignar habitacion')
    context_object_name = 'habitacion'

    def post(self, request, *args, **kwargs):
        huesped = modelos.Huesped.objects.get(
            idhuesped=self.request.POST.get('huesped'))
       
        habitacion = modelos.Habitacion.objects.get(
            idhabitacion=self.request.POST.get('habitacion'))
        dias = huesped.fechahasta - huesped.fechadesde
        total = dias.days * habitacion.precio
        detallereserva = modelos.DetalleReserva.objects.get(huesped=huesped)
        if self.request.method == 'POST':
            if self.request.POST.get('habitacion'):
                estadohabitacion = modelos.Estadohabitacion.objects.get(
                    idestado=2)
                habitacion.estadohabitacion = estadohabitacion
                detallereserva.habitacion = habitacion
                detallereserva.total = total
                detallereserva.save()
                habitacion.save()

        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        empleado = modelos.Empleado.objects.get(usuario = self.request.user.idusuario)
        self.object = form.save(commit=False)
        self.object.empleado = empleado
        self.object = form.save()
        # Llamado a procedimiento almacenado
        with connection.cursor() as cursor:
            cursor.callproc('p_actualizar_habitaciones')
        return super(AsignarHabitacion, self).form_valid(form)

    def test_func(self):
        user = self.request.user
        if user.tipousuario.idtipousuario == 2:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')



# PDF
class FacturaPDF(View):

    def link_callback(self, uri, rel):
        """
        Convert HTML URIs to absolute system paths so xhtml2pdf can access those
        resources
        """
        # use short variable names
        sUrl = settings.STATIC_URL  # Typically /static/
        sRoot = settings.STATIC_ROOT  # Typically /home/userX/project_static/
        mUrl = settings.MEDIA_URL  # Typically /static/media/
        mRoot = settings.MEDIA_ROOT  # Typically /home/userX/project_static/media/

        # convert URIs to absolute system paths
        if uri.startswith(mUrl):
            path = os.path.join(mRoot, uri.replace(mUrl, ""))
        elif uri.startswith(sUrl):
            path = os.path.join(sRoot, uri.replace(sUrl, ""))
        else:
            return uri  # handle absolute uri (ie: http://some.tld/foo.png)

        # make sure that file exists
        if not os.path.isfile(path):
            raise Exception(
                'media URI must start with %s or %s' % (sUrl, mUrl)
            )
        return path

    def get(self, request, *args, **kwargs):
        template = get_template('pdf/factura.html')
        factura = modelos.Factura.objects.get(idfactura=self.kwargs['pk'])
        detallefactura = modelos.Detallefactura.objects.all().filter(factura=factura)
        usuario = self.request.user
        subtotal = 0
        for d in detallefactura:
            subtotal = subtotal + d.total
        iva = round(subtotal * 0.19)

        if factura.cliente.usuario.idusuario == usuario.idusuario or self.request.user.tipousuario.idtipousuario == 2:

            context = {
                'factura': factura,
                'detallefactura': detallefactura,
                'subtotal': subtotal,
                'iva': iva,
                'total': subtotal + iva,
                'logo' : '{}{}'.format(settings.MEDIA_URL, 'logonofondo.png'),
            }
            html = template.render(context)
            response = HttpResponse(content_type="application/pdf")
            response['Content-Disposition'] = 'attachment; filename=' + \
                "factura-numero-" + self.kwargs['pk'] + ".pdf"
            pisaStatus = pisa.CreatePDF(
                html, dest=response,
                link_callback=self.link_callback
            )

            

            if pisaStatus.err:
                return HttpResponse('error')
            return response
        else:
            return redirect('listar facturas emitidas')




# VISTAS SOLICITUD DE COMPRA

class SolicitudCompra(UserPassesTestMixin, SuccessMessageMixin, CreateView):
    model = modelos.SolicitudCompra
    form_class = formularios.SolicitudCompraForm
    template_name = 'dashboard/empleado/solicitudcompra.html'
    success_message = 'Solicitud emitida.'
    success_url = reverse_lazy('nueva solicitud compra')

    def test_func(self):
        user = self.request.user
        if user.tipousuario.idtipousuario == 2:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')

    def get_context_data(self, **kwargs):
        context = super(SolicitudCompra, self).get_context_data(**kwargs)
        context['huespedes'] = modelos.Huesped.objects.all().filter(
            habitacion__isnull=False, fechahasta__gte = date.today()).order_by('idhuesped')
        context['servicios'] = modelos.Serviciocomedor.objects.all()
        return context

    def form_valid(self, form):
        huesped = modelos.Huesped.objects.get(
            idhuesped=self.request.POST.get('huesped'))
        servicio = modelos.Serviciocomedor.objects.get(
            idservicio=self.request.POST.get('servicio'))
        empleado = modelos.Empleado.objects.get(usuario = self.request.user.idusuario)
        self.object = form.save(commit=False)
        self.object.fecha = date.today()
        self.object.huesped = huesped
        self.object.serviciocomedor = servicio
        self.object.empleado = empleado
        self.object = form.save()
        return super(SolicitudCompra, self).form_valid(form)


class ListarClientesFactura(UserPassesTestMixin, SuccessMessageMixin, ListView):
    model = modelos.Factura
    template_name = 'dashboard/empleado/emitirfactura.html'

    def test_func(self):
        user = self.request.user
        if user.tipousuario.idtipousuario == 2:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')

    def get_context_data(self, **kwargs):

        huespedes = modelos.Huesped.objects.values_list(
            'idhuesped', flat=True).filter(habitacion__isnull=False)
        facturas = modelos.Detallefactura.objects.all().values_list('huesped', flat=True)

        facturascompras = modelos.Detallefactura.objects.all(
        ).values_list('solicitudcompra', flat=True)
        solicitudcompraid = modelos.SolicitudCompra.objects.all(
        ).values_list('idsolicitud', flat=True)

        q = huespedes.difference(facturas)
        q2 = solicitudcompraid.difference(facturascompras)
        q3 = modelos.SolicitudCompra.objects.all().values_list(
            'huesped', flat=True).filter(idsolicitud__in=q2)
        idclientes = modelos.Huesped.objects.all().values_list(
            'cliente', flat=True).filter(Q(idhuesped__in=q) | Q(idhuesped__in=q3))
        clientes = modelos.Cliente.objects.all().filter(idcliente__in=idclientes)
        context = super(ListarClientesFactura, self).get_context_data(**kwargs)
        context['clientes'] = clientes
        return context


class EmitirFactura(UserPassesTestMixin, SuccessMessageMixin, CreateView):
    model = modelos.Factura
    form_class = formularios.FacturaForm
    template_name = 'dashboard/empleado/emitirfacturaform.html'
    success_message = 'Factura emitida.'
    success_url = reverse_lazy('emitir factura')

    def test_func(self, **kwargs):
        user = self.request.user
        if user.tipousuario.idtipousuario == 2:
            huespedes = modelos.Huesped.objects.values_list('idhuesped', flat=True).filter(
                habitacion__isnull=False, cliente=self.kwargs['pk'])
            facturas = modelos.Detallefactura.objects.all().values_list('huesped', flat=True)

            q = huespedes.difference(facturas)
            detallereserva = modelos.DetalleReserva.objects.all().filter(huesped__in=q)
            facturassolicitud = modelos.Detallefactura.objects.all(
            ).values_list('solicitudcompra', flat=True)
            solicitudcompraid = modelos.SolicitudCompra.objects.all(
            ).values_list('idsolicitud').filter(huesped__in=huespedes)

            q2 = solicitudcompraid.difference(facturassolicitud)

            solicitudcompra = modelos.SolicitudCompra.objects.all().filter(idsolicitud__in=q2)
            if detallereserva:
                return True
            
            if solicitudcompra:
                return True
            return False

    def handle_no_permission(self):
        user = self.request.user
        if user.tipousuario.idtipousuario != 2:
            return redirect('dashboard')
        return redirect('emitir factura')
        

    def get_context_data(self, **kwargs):
        huespedes = modelos.Huesped.objects.values_list('idhuesped', flat=True).filter(
            habitacion__isnull=False, cliente=self.kwargs['pk'])
        facturas = modelos.Detallefactura.objects.all().values_list('huesped', flat=True)

        q = huespedes.difference(facturas)
        detallereserva = modelos.DetalleReserva.objects.all().filter(huesped__in=q)
        facturassolicitud = modelos.Detallefactura.objects.all(
        ).values_list('solicitudcompra', flat=True)
        solicitudcompraid = modelos.SolicitudCompra.objects.all(
        ).values_list('idsolicitud').filter(huesped__in=huespedes)

        q2 = solicitudcompraid.difference(facturassolicitud)

        solicitudcompra = modelos.SolicitudCompra.objects.all().filter(idsolicitud__in=q2)

        context = super(EmitirFactura, self).get_context_data(**kwargs)
        context['detallesreservas'] = detallereserva
        context['solicitudescompras'] = solicitudcompra
        return context

    def form_valid(self, form, **kwargs):
        reservas = self.request.POST.getlist('chkreserva[]')
        servicios = self.request.POST.getlist('chkservicio[]')
        clienteid = self.kwargs['pk']
        cliente = modelos.Cliente.objects.get(idcliente=clienteid)
        estadofactura = modelos.Estadofactura.objects.get(idestado=1)
        empleado = modelos.Empleado.objects.get(usuario = self.request.user.idusuario)
        self.object = form.save(commit=False)
        self.object.giro = 'Hostal Doña Clarita'
        self.object.fechafactura = date.today()
        self.object.cliente = cliente
        self.object.empleado = empleado
        self.object.estadofactura = estadofactura
        self.object.subtotal = 0
        self.object.iva = 0
        self.object.total = 0
        self.object = form.save()

        for reservaid in reservas:
            reserva = modelos.DetalleReserva.objects.get(idreserva=reservaid)

            detallefact = modelos.Detallefactura()
            detallefact.total = reserva.total
            detallefact.factura = self.object
            detallefact.huesped = reserva.huesped
            detallefact.detallereserva = reserva
            detallefact.save()

        for solicitudid in servicios:
            servicio = modelos.SolicitudCompra.objects.get(
                idsolicitud=solicitudid)

            detallefact = modelos.Detallefactura()
            detallefact.total = servicio.serviciocomedor.precio * servicio.cantidad
            detallefact.factura = self.object
            detallefact.huesped = servicio.huesped
            detallefact.solicitudcompra = servicio
            detallefact.save()

        detfactura = modelos.Detallefactura.objects.all().filter(factura = self.object.idfactura)
        
        subtotal = 0
        for d in detfactura:
            subtotal = subtotal + d.total
        
        iva = round(subtotal * 0.19)
        total = subtotal + iva
        
        self.object.subtotal = subtotal
        self.object.total = total
        self.object.iva = iva
        self.object = form.save()
        return super(EmitirFactura, self).form_valid(form)


class ListarFacturasEmitidas(UserPassesTestMixin, TemplateView):
    template_name = 'dashboard/cliente/listafacturas.html'

    def test_func(self):
        user = self.request.user
        if user.tipousuario.idtipousuario == 3 or user.tipousuario.idtipousuario == 2:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')

    def get_context_data(self, **kwargs):

        if self.request.user.tipousuario.idtipousuario == 3:

            cliente = modelos.Cliente.objects.get(usuario=self.request.user.idusuario)

            facturas = modelos.Factura.objects.all().filter(cliente=cliente.idcliente).order_by('idfactura')

            facturasid = modelos.Factura.objects.values_list('idfactura', flat=True).filter(cliente=cliente.idcliente).order_by('idfactura')

            if self.request.GET.get('estado'):
                get_estado = self.request.GET.get('estado').replace('+',' ').split('?')
                full_path = self.request.get_full_path()
                
                if full_path.find('?page=') == -1:
                    estado = modelos.Estadofactura.objects.get(descripcion = self.request.GET.get('estado'))
                    facturas = modelos.Factura.objects.all().filter(cliente=cliente.idcliente, estadofactura = estado).order_by('idfactura')
                    facturasid = modelos.Factura.objects.values_list('idfactura', flat=True).filter(cliente=cliente.idcliente, estadofactura = estado).order_by('idfactura')
                else:
                    estado = modelos.Estadofactura.objects.get(descripcion = get_estado[0])
                    facturas = modelos.Factura.objects.all().filter(cliente=cliente.idcliente, estadofactura = estado).order_by('idfactura')
                    facturasid = modelos.Factura.objects.values_list('idfactura', flat=True).filter(cliente=cliente.idcliente, estadofactura = estado).order_by('idfactura')
        
        
        if self.request.user.tipousuario.idtipousuario == 2:    

            facturas = modelos.Factura.objects.all().order_by('idfactura')

            facturasid = modelos.Factura.objects.values_list('idfactura', flat=True).order_by('idfactura')

            if self.request.GET.get('estado'):
                get_estado = self.request.GET.get('estado').replace('+',' ').split('?')
                full_path = self.request.get_full_path()
                
                if full_path.find('?page=') == -1:
                    estado = modelos.Estadofactura.objects.get(descripcion = self.request.GET.get('estado'))
                    facturas = modelos.Factura.objects.all().filter(estadofactura = estado).order_by('idfactura')
                    facturasid = modelos.Factura.objects.values_list('idfactura', flat=True).filter(estadofactura = estado).order_by('idfactura')
                else:
                    estado = modelos.Estadofactura.objects.get(descripcion = get_estado[0])
                    facturas = modelos.Factura.objects.all().filter(estadofactura = estado).order_by('idfactura')
                    facturasid = modelos.Factura.objects.values_list('idfactura', flat=True).filter(estadofactura = estado).order_by('idfactura')

        detallefactura = modelos.Detallefactura.objects.all().filter(factura__in=facturasid).order_by('iddetalle')


        #PAGINATION
        paginator = Paginator(facturas, 10) # Muestra 10 por pagina.

        numero_paginas = []

        
        for i in range(1, paginator.num_pages+1):
            numero_paginas.append(i)

        page_number = self.request.GET.get('page')
        
        try:
            page_obj = paginator.get_page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.get_page(1)
        except EmptyPage:
            page_obj = paginator.get_page(paginator.num_pages)




        context = super(ListarFacturasEmitidas,self).get_context_data(**kwargs)

        context['facturas'] = page_obj
        context['detalles'] = detallefactura
        context['paginas'] = numero_paginas
        context['estados'] = modelos.Estadofactura.objects.all().order_by('idestado')
        if self.request.GET.get('estado'):
            context['estado'] = get_estado[0]
        return context


class PagoFactura(UserPassesTestMixin, SuccessMessageMixin, UpdateView, FacturaPDF):

    model = modelos.Factura
    form_class = formularios.FacturaForm
    template_name = 'dashboard/cliente/mediodepago.html'
    success_message = 'Pago confirmado'
    context_object_name = 'factura'

    def test_func(self):
        user = self.request.user
        if user.tipousuario.idtipousuario == 3 or user.tipousuario.idtipousuario == 2:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')

    def post(self, request, *args, **kwargs):
        if self.request.method == 'POST':
            if self.request.user.tipousuario.idtipousuario == 3:
                options = int(self.request.POST.get('options'))

                estadofactura = modelos.Estadofactura.objects.get(
                    idestado=options)
            else:
                estadofactura = modelos.Estadofactura.objects.get(idestado=2)
            factura = modelos.Factura.objects.get(idfactura=self.kwargs['pk'])
            factura.estadofactura = estadofactura
            
            factura.save()

        return super().post(request, *args, **kwargs)

    def form_valid(self,form, **kwargs):
        self.object = form.save(commit=False)
        if self.object.estadofactura.idestado == 2:

            self.object.fechapago = date.today()
            response = FacturaPDF.get(self= self, request = self.request)

            html = render_to_string('pdf/email.html', { 'nombre_cliente': self.object.cliente.usuario.nombre + " " + self.object.cliente.usuario.apellido_paterno })
            pdf = response.content
            email = EmailMultiAlternatives(
                subject ='Factura #'+str(self.object.idfactura),
                body  = html,
                to = [self.object.cliente.usuario.correo],
                )
            email.content_subtype = "html"
            email.attach('factura.pdf', pdf, 'application/pdf')
            email.send()
        self.object = form.save()

        return super(PagoFactura, self).form_valid(form)

    def get_success_url(self, **kwargs):
        idfact = int(self.kwargs['pk'])
        factura = modelos.Factura.objects.get(idfactura=idfact)
        if self.request.user.tipousuario.idtipousuario == 3:
            return reverse_lazy('detalle pago', kwargs={'pk': factura.estadofactura.idestado, 'pk2': factura.idfactura})
        else:
            return reverse_lazy('listar facturas emitidas')


class DetallePago(UserPassesTestMixin, SuccessMessageMixin, TemplateView):
    template_name = 'dashboard/cliente/detalledelpago.html'

    def test_func(self):
        user = self.request.user
        if user.tipousuario.idtipousuario == 3:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')

    def get_context_data(self, **kwargs):
        factura = modelos.Factura.objects.get(
            idfactura=int(self.kwargs['pk2']))
        
        total = factura.total
        
        context = super(DetallePago, self).get_context_data(**kwargs)
        context['option'] = int(self.kwargs['pk'])
        context['total'] = total
        return context



# PEDIDOS

class ListaProveedor(UserPassesTestMixin, TemplateView):
    template_name = 'dashboard/empleado/solicitarprod.html'

    def test_func(self):
        user = self.request.user
        if user.tipousuario.idtipousuario == 2:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')

    def get_context_data(self, **kwargs):

        context = super(ListaProveedor, self).get_context_data(**kwargs)
        context['proveedores'] = modelos.Proveedor.objects.all().order_by('idproveedor')
        return context


class SolicitarProducto(UserPassesTestMixin, SuccessMessageMixin, CreateView):
    model = modelos.Pedido
    form_class = formularios.PedidoForm
    template_name = 'dashboard/empleado/seleccionarprod.html'
    success_message = 'Productos solicitados.'
    success_url = reverse_lazy('listar proveedor')

    def test_func(self):
        user = self.request.user
        if user.tipousuario.idtipousuario == 2:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')

    def get_context_data(self, **kwargs):
        productos = modelos.Producto.objects.all().filter(
            proveedor=self.kwargs['pk'])

        context = super(SolicitarProducto, self).get_context_data(**kwargs)
        context['productos'] = productos
        return context

    def form_valid(self, form, **kwargs):
        cantidad = self.request.POST.getlist('cantidad[]')
        idprod = self.request.POST.getlist('idpro[]')
        arr = []
        cont = 0
        c = 0
        for i in cantidad:
            c+=int(i)
        
        if c == 0:
            messages.error(self.request,'Debe solicitar almenos 1 producto.')
            return HttpResponseRedirect(reverse('solicitar producto', kwargs={'pk': self.kwargs['pk']}))
        print(idprod)
        for c in idprod:
            arr.append([c, cantidad[cont]])
            cont = cont+1

        proveedorid = self.kwargs['pk']
        emp = modelos.Empleado.objects.get(usuario=self.request.user.idusuario)

        estadopedido = modelos.Estadopedido.objects.get(idestado=1)

        self.object = form.save(commit=False)
        self.object.observaciones = self.request.POST.get('observaciones')
        self.object.fechapedido = date.today()
        self.object.empleado = emp
        self.object.subtotal = 0
        self.object.iva = 0
        self.object.total = 0
        self.object.estadopedido = estadopedido
        self.object.proveedor = modelos.Proveedor.objects.get(
            idproveedor=proveedorid)
        self.object = form.save()

        for pedido in arr:
            if int(pedido[1]) > 0:
                producto = modelos.Producto.objects.get(
                    idproducto=int(pedido[0]))
                detallepedido = modelos.Detallepedido()
                detallepedido.cantidad = int(pedido[1])
                detallepedido.total = int(pedido[1])*producto.precio
                detallepedido.pedido = self.object
                detallepedido.producto = producto
                detallepedido.estado = modelos.EstadoDetallePedido.objects.get(idestado = 3)
                detallepedido.save()
        
        dp = modelos.Detallepedido.objects.all().filter(pedido = self.object.idpedido)
        subtotal = 0
        iva = 0
        total = 0
        for detalle in dp:
            subtotal += detalle.total
        
        iva = round(subtotal*0.19)
        total = subtotal + iva

        self.object = form.save(commit=False)
        self.object.subtotal = subtotal
        self.object.iva = iva
        self.object.total = total
        self.object = form.save()
        return super(SolicitarProducto, self).form_valid(form)


class ListarPedidos(UserPassesTestMixin, SuccessMessageMixin, ListView):

    model = modelos.Pedido
    template_name = 'dashboard/empleado/solicitudes.html'
    context_object_name = 'pedido'

    def test_func(self):
        user = self.request.user
        if user.tipousuario.idtipousuario == 2 or user.tipousuario.idtipousuario == 4:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')

    def get_context_data(self, **kwargs):
        if self.request.user.tipousuario.idtipousuario == 2:

            pedidos = modelos.Pedido.objects.all().order_by('idpedido')

            if self.request.GET.get('estado'):
                get_estado = self.request.GET.get('estado').replace('+',' ').split('?')
                full_path = self.request.get_full_path()
                
                if full_path.find('?page=') == -1:
                    estado = modelos.Estadopedido.objects.get(descripcion = self.request.GET.get('estado'))
                    pedidos = modelos.Pedido.objects.all().filter(estadopedido = estado).order_by('idpedido')
                else:
                    estado = modelos.Estadopedido.objects.get(descripcion = get_estado[0])
                    pedidos = modelos.Pedido.objects.all().filter(estadopedido = estado).order_by('idpedido')
                    
        
        if self.request.user.tipousuario.idtipousuario == 4:
            proveedor = modelos.Proveedor.objects.get(
                usuario=self.request.user.idusuario)
            pedidos = modelos.Pedido.objects.all().filter(proveedor=proveedor.idproveedor).order_by('idpedido')
            if  self.request.GET.get('estado'):
                get_estado = self.request.GET.get('estado').replace('+',' ').split('?')
                full_path = self.request.get_full_path()

                if full_path.find('?page=') == -1:

                    
                    estado = modelos.Estadopedido.objects.get(descripcion = self.request.GET.get('estado'))
                    pedidos = modelos.Pedido.objects.all().filter(proveedor=proveedor.idproveedor,estadopedido = estado).order_by('idpedido')
                    

                else:

                    
                    estado = modelos.Estadopedido.objects.get(descripcion = get_estado[0])
                    pedidos = modelos.Pedido.objects.all().filter(proveedor=proveedor.idproveedor, estadopedido = estado).order_by('idpedido')
                    




        #PAGINATION
        paginator = Paginator(pedidos, 10) # Muestra 10 por pagina.

        numero_paginas = []

        
        for i in range(1, paginator.num_pages+1):
            numero_paginas.append(i)

        
        page_number = self.request.GET.get('page')
        

        


        
        
        try:
            page_obj = paginator.get_page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.get_page(1)
        except EmptyPage:
            page_obj = paginator.get_page(paginator.num_pages)


        

        context = super(ListarPedidos, self).get_context_data(**kwargs)
        context['pedidos'] = page_obj
        context['detalles'] = modelos.Detallepedido.objects.all()
        context['numero_paginas'] = numero_paginas
        context['estados'] = modelos.Estadopedido.objects.all()
        if self.request.GET.get('estado'):
            context['estado'] = get_estado[0]
            
        return context


class AdministrarSolicitud(UserPassesTestMixin, SuccessMessageMixin, UpdateView):

    model = modelos.Pedido
    form_class = formularios.PedidoForm
    template_name = 'dashboard/empleado/solicitudes.html'
    success_message = 'Solicitud Actualizada'
    context_object_name = 'pedido'
    success_url = reverse_lazy('listar pedidos')

    def test_func(self):
        user = self.request.user
        if user.tipousuario.idtipousuario == 2 or user.tipousuario.idtipousuario == 4:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')

    def post(self, request, *args, **kwargs):

        if self.request.method == 'POST':
            pedido = modelos.Pedido.objects.get(idpedido=self.kwargs['pk'])
            if 'rechazar' in self.request.POST:
                estadopedido = modelos.Estadopedido.objects.get(idestado=4)

            if 'aceptar' in self.request.POST and self.request.user.tipousuario.idtipousuario == 4:
                estadopedido = modelos.Estadopedido.objects.get(idestado=2)
                

            if 'recibir' in self.request.POST and self.request.user.tipousuario.idtipousuario == 2:

                aceptados = self.request.POST.getlist('chkAceptar[]')
                
                for i in aceptados:
                    detalle = modelos.Detallepedido.objects.get(idedetalle = i)
                    detalle.estado = modelos.EstadoDetallePedido.objects.get(idestado = 1)
                    detalle.save()
                
                
                rechazados = modelos.Detallepedido.objects.all().filter(pedido=pedido.idpedido, estado = 3)
                
                if rechazados:
                    for j in rechazados:
                        detalle_re = modelos.Detallepedido.objects.get(idedetalle = j.idedetalle)
                        detalle_re.estado = modelos.EstadoDetallePedido.objects.get(idestado = 2)
                        detalle_re.save()

                detallepedido = modelos.Detallepedido.objects.all().filter(pedido=pedido.idpedido, estado = 1)
                pedido.fechaentrega = date.today()
                
                for dp in detallepedido:
                    producto = modelos.Producto.objects.get(
                        idproducto=dp.producto.idproducto)
                    producto.stock = producto.stock + dp.cantidad
                    producto.save()
                    
                    if(len(modelos.Recepcionproducto.objects.all().filter(detallepedido = dp.idedetalle)) == 0):

                        recepcion = modelos.Recepcionproducto()
                        recepcion.codigo = self.codigo_producto(dp.idedetalle, self.request.POST.get('date'+str(dp.idedetalle)))
                        recepcion.fecharecepcion = date.today()
                        recepcion.detallepedido = modelos.Detallepedido.objects.get(idedetalle = dp.idedetalle)
                        recepcion.empleado = modelos.Empleado.objects.get(usuario = self.request.user.idusuario)
                        recepcion.save()

                if len(modelos.Detallepedido.objects.all().filter(pedido=pedido.idpedido, estado = 2)) == 0:

                    estadopedido = modelos.Estadopedido.objects.get(idestado=3)
                    detalles = modelos.Detallepedido.objects.all().filter(pedido=pedido.idpedido, estado = 1)
                    subtotal = 0
                    for d in detalles:
                        subtotal += d.total
                    pedido.subtotal = subtotal
                    pedido.iva = round(subtotal*0.19)
                    pedido.total = subtotal + round(subtotal*0.19)
                else:
                    pendientes = modelos.Detallepedido.objects.all().filter(pedido=pedido.idpedido, estado = 2)
                    subtotal = 0
                    for p in pendientes:
                        subtotal += p.total
                
                    estadopedido = modelos.Estadopedido.objects.get(idestado=1)
                    pedido.subtotal = subtotal
                    pedido.iva = round(subtotal*0.19)
                    pedido.total = subtotal + round(subtotal*0.19)
            pedido.estadopedido = estadopedido
            pedido.save()
        return super().post(request, *args, **kwargs)

    def codigo_producto(self,iddetalle, fechavencimiento):
        detalle = modelos.Detallepedido.objects.get(idedetalle = iddetalle)
        idproveedor = str(detalle.pedido.proveedor.idproveedor).zfill(3)
       
        idproducto = str(detalle.producto.idproducto).zfill(3)
       
        fechavencimiento_numero = '00000000'
        if fechavencimiento:
            fechavencimiento_numero = fechavencimiento.replace("-", "")
            print(fechavencimiento_numero)
            
        idtipoproducto = str(detalle.producto.tipoproducto.idtipo).zfill(3)
        
        return idproveedor[-3:]+idproducto[-3:]+fechavencimiento_numero+idtipoproducto[-3:]


# ACTUALIZAR ESTADO HABITACIONES

class ActualizarEstHab(UserPassesTestMixin, SuccessMessageMixin, FormView):
    form_class = formularios.HabitacionForm
    template_name = 'dashboard/empleado/asignarhabi.html'
    success_message = 'Se actualizo el estado de las habitaciones.'
    success_url = reverse_lazy('asignar habitacion')

    def test_func(self):
        user = self.request.user
        if user.tipousuario.idtipousuario == 2 or user.tipousuario.idtipousuario == 4:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')

    
    def form_valid(self, form):

        form.actualizar_estado()
        return super(ActualizarEstHab, self).form_valid(form)


# INFORMES Y ESTADISTICAS

def obtener_datos(request, *args, **kwargs):
    
    #USUARIO CLIENTE
    if request.user.tipousuario.idtipousuario == 3:

        #DATOS DEL USUARIO
        cliente = modelos.Cliente.objects.get(usuario = request.user.idusuario)
        tipousuario = request.user.tipousuario.idtipousuario
        #ESTADISTICAS RESERVA
        #TOTAL RESERVAS FINALIZADAS
        reservasfinalizadas = len(modelos.Huesped.objects.all().filter(cliente = cliente.idcliente, fechahasta__lte = date.today()))
        #TOTALRESERVAS ACTIVAS
        reservasactivas = len(modelos.Huesped.objects.all().filter(cliente = cliente.idcliente, fechahasta__gte = date.today()))
        total_reservas = reservasfinalizadas + reservasactivas
        # ARRAYS CON LOS DATOS Y LABELS
        data_reserva = []
        data_reserva.append(reservasfinalizadas)
        data_reserva.append(reservasactivas)
        labels_reserva = []
        labels_reserva.append("Reservas Finalizadas")
        labels_reserva.append("Reservas Activas")

        #ESTADISTICAS FACTURA
        labels_factura = []
        labels_factura.append("Pagadas")
        labels_factura.append("En Proceso")
        labels_factura.append("No Pagadas")
        data_factura = []
        # TOTAL DE FACTURAS CON PAGOS CONFIRMADOS
        facturas_pagadas = len(modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 2))
        # TOTAL FACTURAS EMITIDAS Y NO PAGADAS
        facturas_no_pagadas = len(modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 1))
        # TOTAL FACTURAS EN PROCESO DE PAGO       
        facturas_en_proceso = len(modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura__in = (3,4)))
        data_factura.append(facturas_pagadas)
        data_factura.append(facturas_en_proceso)
        data_factura.append(facturas_no_pagadas)
        # FACTURAS TOTALES
        facturas_totales = facturas_no_pagadas + facturas_en_proceso + facturas_pagadas
        
        # FACTURA TOTAL MES (FTM)
        ftm_labels = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        ftm_data = []

        # OTRAS ESTADISTICAS
        huespedes_totales = len(modelos.Huesped.objects.all().filter(cliente = cliente.idcliente))

        facturas_ids = modelos.Factura.objects.values_list('idfactura', flat=True).filter(cliente = cliente.idcliente)
        servicios_totales = len(modelos.Detallefactura.objects.all().filter(factura__in = facturas_ids))
        
        habitaciones_disponibles = len(modelos.Habitacion.objects.all().filter(estadohabitacion = 1))
        
        for i in range(1,13):
            facturas = modelos.Factura.objects.all().filter(fechapago__year = date.today().strftime("%Y"),fechapago__month = i, cliente = cliente.idcliente)
            ftm_total = 0
            for factura in facturas:
                ftm_total = ftm_total + factura.total
            
            ftm_data.append(ftm_total)

        
        # JSON    
        data = {
            
                "labels_reserva" : labels_reserva,
                "data_reserva": data_reserva,
                "tipousuario": tipousuario,
                "total_reserva": total_reservas,
                "labels_factura" : labels_factura,
                "data_factura" : data_factura,
                "total_factura" : facturas_totales,
                "ftm_labels" : ftm_labels,
                "ftm_data" : ftm_data,
                "huespedes" : huespedes_totales,
                "servicios" : servicios_totales,
                "habitaciones_disponible" : habitaciones_disponibles
             
        }
        return JsonResponse(data)

    if request.user.tipousuario.idtipousuario == 2:
        #DATOS DEL USUARIO
        empleado = modelos.Empleado.objects.get(usuario = request.user.idusuario)
        tipousuario = request.user.tipousuario.idtipousuario

        # ESTADISTICAS NO GRAFICOS
        habitaciones_asignadas = len(modelos.Huesped.objects.all().filter(empleado = empleado.idempleado))
        servicios_solicitados = len(modelos.SolicitudCompra.objects.all().filter(empleado = empleado.idempleado))
        facturas_emitidas = len(modelos.Factura.objects.all().filter(empleado = empleado.idempleado))
        ganancias_totales = 0
        facturas_gt = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 2)
        for f in facturas_gt:
            ganancias_totales += f.total
        
        # GRAFICO GANANCIAS
        ganancias_labels = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        ganancias_data = []

        for i in range(1,13):
            ganancia = 0
            fact = modelos.Factura.objects.all().filter(fechapago__year = date.today().strftime("%Y"),fechapago__month = i, empleado = empleado.idempleado)
            for j in fact:
                ganancia += j.total
            ganancias_data.append(ganancia)

        
        # SOLICITUDES PRODUCTOS
        sdp_labels = ["Solicitados", "En Transito", "Recibidos", "Rechazaados"]
        sdp_data = []

        sdp_solicitados = len(modelos.Pedido.objects.all().filter(estadopedido = 1))
        sdp_entransito = len(modelos.Pedido.objects.all().filter(estadopedido = 2))
        sdp_recibidos = len(modelos.Pedido.objects.all().filter(estadopedido = 3))
        sdp_rechazados = len(modelos.Pedido.objects.all().filter(estadopedido = 4))
        sdp_total = sdp_solicitados + sdp_entransito + sdp_recibidos + sdp_rechazados

        sdp_data.append(sdp_solicitados)
        sdp_data.append(sdp_entransito)
        sdp_data.append(sdp_recibidos)
        sdp_data.append(sdp_rechazados)

        # FACTURAS
        f_labels = ["Pagadas","En Proceso","No Pagadas"]
        f_data = []

        f_pagadas = len(modelos.Factura.objects.all().filter(estadofactura = 2))
        f_enproceso = len(modelos.Factura.objects.all().filter(estadofactura__in = (3,4)))
        f_nopagada = len(modelos.Factura.objects.all().filter(estadofactura = 1))

        f_data.append(f_pagadas)
        f_data.append(f_enproceso)
        f_data.append(f_nopagada)

        f_total = f_pagadas + f_enproceso + f_nopagada


        # JSON    
        data = {
            
                "tipousuario": tipousuario,
                "habitaciones_asignadas" : habitaciones_asignadas,
                "servicios_solicitados" : servicios_solicitados,
                "facturas_emitidas" : facturas_emitidas,
                "ganacias_totales" : ganancias_totales,
                "ganancias_labels" : ganancias_labels,
                "ganancias_data" : ganancias_data,
                "sdp_labels" : sdp_labels,
                "sdp_data" : sdp_data,
                "sdp_total" : sdp_total,
                "f_labels" : f_labels,
                "f_data" : f_data,
                "f_total" : f_total

               
             
        }
        return JsonResponse(data)

    if request.user.tipousuario.idtipousuario == 4:
        #DATOS DEL USUARIO
        proveedor = modelos.Proveedor.objects.get(usuario = request.user.idusuario)
        tipousuario = request.user.tipousuario.idtipousuario

        # ESTADISTICAS NO GRAFICOS
        pedidos_solicitados = len(modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 1))
        pedidos_transito = len(modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 2))
        pedidos_entregados = len(modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 3))
        pedidos_rechazados = len(modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 4))
        
        
        # GRAFICO GANANCIAS
        ganancias_labels = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        ganancias_data = []

        for i in range(1,13):
            ganancia = 0
            pedido = modelos.Pedido.objects.all().filter(fechaentrega__year = date.today().strftime("%Y"),fechaentrega__month = i, proveedor = proveedor.idproveedor)
            for j in pedido:
                ganancia += j.total
            ganancias_data.append(ganancia)

        # JSON    
        data = {
            
                "pedidos_solicitados": pedidos_solicitados,
                "pedidos_transito" : pedidos_transito,
                "pedidos_entregados" : pedidos_entregados,
                "pedidos_rechazados" : pedidos_rechazados,
                "tipousuario" : tipousuario,
                "ganancias_data" : ganancias_data,
                "ganancias_labels" : ganancias_labels
             
        }
        return JsonResponse(data)

    if request.user.tipousuario.idtipousuario == 1:
        usuarios = len(modelos.Usuario.objects.all())
        productos = len(modelos.Producto.objects.all())
        servicios = len(modelos.Serviciocomedor.objects.all())
        habitaciones = len(modelos.Habitacion.objects.all())

        # GRAFICO GANANCIAS
        ganancias_labels = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        ganancias_data = []

        for i in range(1,13):
            ganancia = 0
            fact = modelos.Factura.objects.all().filter(fechapago__year = date.today().strftime("%Y"),fechapago__month = i)
            for j in fact:
                ganancia += j.total
            ganancias_data.append(ganancia)

        # SOLICITUDES PRODUCTOS
        sdp_labels = ["Solicitados", "En Transito", "Recibidos", "Rechazaados"]
        sdp_data = []

        sdp_solicitados = len(modelos.Pedido.objects.all().filter(estadopedido = 1))
        sdp_entransito = len(modelos.Pedido.objects.all().filter(estadopedido = 2))
        sdp_recibidos = len(modelos.Pedido.objects.all().filter(estadopedido = 3))
        sdp_rechazados = len(modelos.Pedido.objects.all().filter(estadopedido = 4))
        sdp_total = sdp_solicitados + sdp_entransito + sdp_recibidos + sdp_rechazados

        sdp_data.append(sdp_solicitados)
        sdp_data.append(sdp_entransito)
        sdp_data.append(sdp_recibidos)
        sdp_data.append(sdp_rechazados)


        # FACTURAS
        f_labels = ["Pagadas","En Proceso","No Pagadas"]
        f_data = []

        f_pagadas = len(modelos.Factura.objects.all().filter(estadofactura = 2))
        f_enproceso = len(modelos.Factura.objects.all().filter(estadofactura__in = (3,4)))
        f_nopagada = len(modelos.Factura.objects.all().filter(estadofactura = 1))

        f_data.append(f_pagadas)
        f_data.append(f_enproceso)
        f_data.append(f_nopagada)

        f_total = f_pagadas + f_enproceso + f_nopagada

        # JSON    
        data = {
            
                "usuarios" : usuarios,
                "productos" : productos,
                "servicios" : servicios,
                "habitaciones" : habitaciones,
                "tipousuario" : 1,
                "ganancias_labels" : ganancias_labels,
                "ganancias_data" : ganancias_data,
                "sdp_labels" : sdp_labels,
                "sdp_data" : sdp_data,
                "sdp_total" : sdp_total,
                "f_labels" : f_labels,
                "f_data" : f_data,
                "f_total" : f_total
                
        }
        return JsonResponse(data)


    return HttpResponse("NO DATA")

class Informes(UserPassesTestMixin, TemplateView):
    template_name = 'dashboard/informes/informesview.html'

    def test_func(self):
        user = self.request.user
        if user:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')

    def post(self, request, *args, **kwargs):
        
        opciones = self.request.POST.getlist('opcion[]')

        if opciones:
            tipousuario = self.request.user.tipousuario.idtipousuario
            wb = Workbook()
            bandera = True
            
            if tipousuario == 2 or tipousuario == 1:
                if tipousuario == 2:
                    empleado = modelos.Empleado.objects.get(usuario = self.request.user.idusuario)
                
                rep = True
                rep_pedidos = True
                for op in opciones:
                
                    if int(op) == 1:

                       
                        if bandera:
                            ws = wb.active
                            ws.title = "Informe Habitaciones Asignadas"
                            bandera = False
                        else:
                            ws = wb.create_sheet("Informe Habitaciones Asignadas")

                        # TITULO DEL REPORTE
                        ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
                        ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                        ws['B1'].fill = PatternFill(start_color = '3D3D3D', end_color = '3D3D3D', fill_type = "solid")
                        ws['B1'].font = Font(name = 'Calibri', size = 12, bold = True, color = 'FFFFFF')

                        ws['B1'] = "INFORME DE HABITACIONES ASIGNADAS"
                        ws.row_dimensions[1].height = 45
                        ws.merge_cells('B1:H1')
                    
                        # TAMAÑO COLUMNA
                        ws.column_dimensions['B'].width = 15
                        ws.column_dimensions['C'].width = 22
                        ws.column_dimensions['D'].width = 22
                        ws.column_dimensions['E'].width = 22
                        ws.column_dimensions['F'].width = 22
                        ws.column_dimensions['G'].width = 22
                        ws.column_dimensions['H'].width = 22


                        # CAMPOS
                        ws['B3'] = "RUT"
                        ws['C3'] = "NOMBRE"
                        ws['D3'] = "APELLIDO PATERNO"
                        ws['E3'] = "APELLIDO MATERNO"
                        ws['F3'] = "FECHA DESDE"
                        ws['G3'] = "FECHA HASTA"
                        ws['H3'] = "NUMERO HABITACION"


                        # ESTILO CABEZERA 
                        for i in range(2, 9):
                            ws.cell(row = 3, column = i).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = 3, column = i).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = 3, column = i).fill = PatternFill(start_color = 'A0A0A0', end_color = 'A0A0A0', fill_type = "solid")
                            ws.cell(row = 3, column = i).font = Font(name = 'Calibri', size = 12, bold = True)

                        
                        # DATOS
                        cont = 4
                        if tipousuario == 2:

                            habitaciones_asignadas = modelos.Huesped.objects.all().filter(empleado = empleado.idempleado, habitacion__isnull = False).order_by('fechadesde')
                        
                            if self.request.POST.get('fechadesdeHuespedes'):
                                habitaciones_asignadas = modelos.Huesped.objects.all().filter(empleado = empleado.idempleado,habitacion__isnull = False,  fechadesde__gte = self.request.POST.get('fechadesdeHuespedes')).order_by('fechadesde')

                            if self.request.POST.get('fechahastaHuespedes'):
                                habitaciones_asignadas = modelos.Huesped.objects.all().filter(empleado = empleado.idempleado,habitacion__isnull = False, fechahasta__lte = self.request.POST.get('fechahastaHuespedes')).order_by('fechadesde')

                        
                            if self.request.POST.get('fechadesdeHuespedes') and self.request.POST.get('fechahastaHuespedes'):
                                habitaciones_asignadas = modelos.Huesped.objects.all().filter(empleado = empleado.idempleado,habitacion__isnull = False, fechadesde__gte = self.request.POST.get('fechadesdeHuespedes'),fechahasta__lte = self.request.POST.get('fechahastaHuespedes')).order_by('fechadesde')
                        
                        if tipousuario == 1:

                            habitaciones_asignadas = modelos.Huesped.objects.all().filter(habitacion__isnull = False).order_by('fechadesde')
                        
                            if self.request.POST.get('fechadesdeHuespedes'):
                                habitaciones_asignadas = modelos.Huesped.objects.all().filter(habitacion__isnull = False,  fechadesde__gte = self.request.POST.get('fechadesdeHuespedes')).order_by('fechadesde')

                            if self.request.POST.get('fechahastaHuespedes'):
                                habitaciones_asignadas = modelos.Huesped.objects.all().filter(habitacion__isnull = False, fechahasta__lte = self.request.POST.get('fechahastaHuespedes')).order_by('fechadesde')

                        
                            if self.request.POST.get('fechadesdeHuespedes') and self.request.POST.get('fechahastaHuespedes'):
                                habitaciones_asignadas = modelos.Huesped.objects.all().filter(habitacion__isnull = False, fechadesde__gte = self.request.POST.get('fechadesdeHuespedes'),fechahasta__lte = self.request.POST.get('fechahastaHuespedes')).order_by('fechadesde')
                        




                        for ha in habitaciones_asignadas:
                            ws.cell(row = cont, column = 2).value = ha.rut
                            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 3).value = ha.nombre
                            ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 4).value = ha.apellidopaterno
                            ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 5).value = ha.apellidomaterno
                            ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 6).value = ha.fechadesde
                            ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 7).value = ha.fechahasta
                            ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 8).value = ha.habitacion.numero
                            ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            cont += 1

                    if int(op) == 2:

                        

                        if bandera:
                            ws = wb.active
                            ws.title = "Informe Compras Realizadas"
                            bandera = False
                        else:
                            ws = wb.create_sheet("Informe Compras Realizadas")

                        # TITULO DEL REPORTE
                        ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
                        ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                        ws['B1'].fill = PatternFill(start_color = '3D3D3D', end_color = '3D3D3D', fill_type = "solid")
                        ws['B1'].font = Font(name = 'Calibri', size = 12, bold = True, color = 'FFFFFF')

                        ws['B1'] = "INFORME DE COMPRAS REALIZADAS"
                        ws.row_dimensions[1].height = 45
                        ws.merge_cells('B1:L1')

                        # TAMAÑO COLUMNA
                        ws.column_dimensions['B'].width = 15
                        ws.column_dimensions['C'].width = 22
                        ws.column_dimensions['D'].width = 22
                        ws.column_dimensions['E'].width = 22
                        ws.column_dimensions['F'].width = 22
                        ws.column_dimensions['G'].width = 22
                        ws.column_dimensions['H'].width = 40
                        ws.column_dimensions['I'].width = 22
                        ws.column_dimensions['J'].width = 15
                        ws.column_dimensions['K'].width = 15
                        ws.column_dimensions['L'].width = 15

                        # CAMPOS
                        ws['B3'] = "RUT"
                        ws['C3'] = "EMPRESA"
                        ws['D3'] = "NOMBRE"
                        ws['E3'] = "APELLIDO PATERNO"
                        ws['F3'] = "APELLIDO MATERNO"
                        ws['G3'] = "FECHA COMPRA"
                        ws['H3'] = "SERVICIO"
                        ws['I3'] = "CANTIDAD"
                        ws['J3'] = "SUBTOTAL"
                        ws['K3'] = "IVA"
                        ws['L3'] = "TOTAL"

                        # ESTILO CABEZERA 
                        for i in range(2, 13):
                            ws.cell(row = 3, column = i).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = 3, column = i).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = 3, column = i).fill = PatternFill(start_color = 'A0A0A0', end_color = 'A0A0A0', fill_type = "solid")
                            ws.cell(row = 3, column = i).font = Font(name = 'Calibri', size = 12, bold = True)

                        
                        # DATOS
                        cont = 4

                        if tipousuario == 2:
                            factura = modelos.Factura.objects.all().values_list('idfactura', flat=True).filter(empleado = empleado.idempleado, estadofactura = 2).order_by('fechapago')
                        
                            if self.request.POST.get('fechadesdeHuespedes'):
                                factura = modelos.Factura.objects.all().values_list('idfactura', flat=True).filter(empleado = empleado.idempleado, estadofactura = 2, fechapago__gte = self.request.POST.get('fechadesdeHuespedes')).order_by('fechapago')
                                

                            if self.request.POST.get('fechahastaHuespedes'):
                                factura = modelos.Factura.objects.all().values_list('idfactura', flat=True).filter(empleado = empleado.idempleado, estadofactura = 2, fechapago__lte = self.request.POST.get('fechahastaHuespedes')).order_by('fechapago')
                                
                            
                            if self.request.POST.get('fechadesdeHuespedes') and self.request.POST.get('fechahastaHuespedes'):
                                factura = modelos.Factura.objects.all().values_list('idfactura', flat=True).filter(empleado = empleado.idempleado, estadofactura = 2, fechapago__gte = self.request.POST.get('fechadesdeHuespedes'), fechapago__lte = self.request.POST.get('fechahastaHuespedes')).order_by('fechapago')
                                
                           
                        if tipousuario == 1:
                            factura = modelos.Factura.objects.all().values_list('idfactura', flat=True).filter(estadofactura = 2).order_by('fechapago')
                        
                            if self.request.POST.get('fechadesdeHuespedes'):
                                factura = modelos.Factura.objects.all().values_list('idfactura', flat=True).filter(estadofactura = 2, fechapago__gte = self.request.POST.get('fechadesdeHuespedes')).order_by('fechapago')
                                

                            if self.request.POST.get('fechahastaHuespedes'):
                                factura = modelos.Factura.objects.all().values_list('idfactura', flat=True).filter(estadofactura = 2, fechapago__lte = self.request.POST.get('fechahastaHuespedes')).order_by('fechapago')
                                
                            
                            if self.request.POST.get('fechadesdeHuespedes') and self.request.POST.get('fechahastaHuespedes'):
                                factura = modelos.Factura.objects.all().values_list('idfactura', flat=True).filter(estadofactura = 2, fechapago__gte = self.request.POST.get('fechadesdeHuespedes'), fechapago__lte = self.request.POST.get('fechahastaHuespedes')).order_by('fechapago')
                                
                        
                        compras_realizadas = modelos.Detallefactura.objects.all().filter(factura__in = factura, solicitudcompra__isnull = False).order_by('factura')

                        for cr in compras_realizadas:
                            ws.cell(row = cont, column = 2).value = cr.huesped.rut
                            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 3).value = cr.huesped.cliente.nombre
                            ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))                            
                            ws.cell(row = cont, column = 4).value = cr.huesped.nombre
                            ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 5).value = cr.huesped.apellidopaterno
                            ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 6).value = cr.huesped.apellidomaterno
                            ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 7).value = cr.factura.fechapago
                            ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 8).value = cr.solicitudcompra.serviciocomedor.plato
                            ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 9).value = cr.solicitudcompra.cantidad
                            ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 10).value = cr.total
                            ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 11).value = round(cr.total*0.19)
                            ws.cell(row = cont, column = 11).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 11).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 12).value = cr.total + round(cr.total*0.19)
                            ws.cell(row = cont, column = 12).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 12).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            cont += 1

                    if int(op) == 3 or int(op) == 4 or int(op) ==5 or int(op) == 6:
                        if rep:
                            rep = False
                            if bandera:
                                ws = wb.active
                                ws.title = "Informe Facturas"
                                bandera = False
                            else:
                                ws = wb.create_sheet("Informe Facturas")
                            
                            # TITULO DEL REPORTE
                            ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
                            ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws['B1'].fill = PatternFill(start_color = '3D3D3D', end_color = '3D3D3D', fill_type = "solid")
                            ws['B1'].font = Font(name = 'Calibri', size = 12, bold = True, color = 'FFFFFF')

                            ws['B1'] = "INFORME DE FACTURAS"
                            ws.row_dimensions[1].height = 45
                            ws.merge_cells('B1:J1')

                            # TAMAÑO COLUMNAS
                            ws.column_dimensions['B'].width = 15
                            ws.column_dimensions['C'].width = 22
                            ws.column_dimensions['D'].width = 22
                            ws.column_dimensions['E'].width = 22
                            ws.column_dimensions['F'].width = 22
                            ws.column_dimensions['G'].width = 22
                            ws.column_dimensions['H'].width = 15
                            ws.column_dimensions['I'].width = 15
                            ws.column_dimensions['J'].width = 15

                            # CAMPOS
                            ws['B3'] = "#"
                            ws['C3'] = "CLIENTE"
                            ws['D3'] = "GIRO"
                            ws['E3'] = "FECHA EMITIDA"
                            ws['F3'] = "FECHA PAGO"
                            ws['G3'] = "ESTADO"
                            ws['H3'] = "SUBTOTAL"
                            ws['I3'] = "IVA"
                            ws['J3'] = "TOTAL"

                            # ESTILO CABEZERA
                            for i in range(2, 11):
                                ws.cell(row = 3, column = i).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = 3, column = i).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = 3, column = i).fill = PatternFill(start_color = 'A0A0A0', end_color = 'A0A0A0', fill_type = "solid")
                                ws.cell(row = 3, column = i).font = Font(name = 'Calibri', size = 12, bold = True)
                            
                            # DATOS
                            cont = 4

                        if tipousuario == 2:

                            facturas_pagadas = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 2).order_by('fechafactura')
                            facturas_no_pagadas = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 1).order_by('fechafactura')
                            facturas_trans_bancaria = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 3).order_by('fechafactura')
                            factura_presencial = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 4).order_by('fechafactura')

                            if self.request.POST.get('fechadesdeFacturas'):
                                facturas_pagadas = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 2, fechafactura__gte = self.request.POST.get('fechadesdeFacturas')).order_by('fechafactura')
                                facturas_no_pagadas = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 1, fechafactura__gte = self.request.POST.get('fechadesdeFacturas')).order_by('fechafactura')
                                facturas_trans_bancaria = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 3, fechafactura__gte = self.request.POST.get('fechadesdeFacturas')).order_by('fechafactura')
                                factura_presencial = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 4, fechafactura__gte = self.request.POST.get('fechadesdeFacturas')).order_by('fechafactura')
                            
                            if self.request.POST.get('fechahastaFacturas'):
                                facturas_pagadas = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 2, fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                                facturas_no_pagadas = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 1, fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                                facturas_trans_bancaria = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 3, fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                                factura_presencial = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 4, fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                            
                            if self.request.POST.get('fechadesdeFacturas') and self.request.POST.get('fechahastaFacturas'):
                                facturas_pagadas = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 2, fechafactura__gte = self.request.POST.get('fechadesdeFacturas'),fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                                facturas_no_pagadas = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 1, fechafactura__gte = self.request.POST.get('fechadesdeFacturas'),fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                                facturas_trans_bancaria = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 3, fechafactura__gte = self.request.POST.get('fechadesdeFacturas'),fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                                factura_presencial = modelos.Factura.objects.all().filter(empleado = empleado.idempleado, estadofactura = 4, fechafactura__gte = self.request.POST.get('fechadesdeFacturas'),fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                                           
                        if tipousuario == 1:

                            facturas_pagadas = modelos.Factura.objects.all().filter(estadofactura = 2).order_by('fechafactura')
                            facturas_no_pagadas = modelos.Factura.objects.all().filter(estadofactura = 1).order_by('fechafactura')
                            facturas_trans_bancaria = modelos.Factura.objects.all().filter(estadofactura = 3).order_by('fechafactura')
                            factura_presencial = modelos.Factura.objects.all().filter(estadofactura = 4).order_by('fechafactura')

                            if self.request.POST.get('fechadesdeFacturas'):
                                facturas_pagadas = modelos.Factura.objects.all().filter(estadofactura = 2, fechafactura__gte = self.request.POST.get('fechadesdeFacturas')).order_by('fechafactura')
                                facturas_no_pagadas = modelos.Factura.objects.all().filter(estadofactura = 1, fechafactura__gte = self.request.POST.get('fechadesdeFacturas')).order_by('fechafactura')
                                facturas_trans_bancaria = modelos.Factura.objects.all().filter(estadofactura = 3, fechafactura__gte = self.request.POST.get('fechadesdeFacturas')).order_by('fechafactura')
                                factura_presencial = modelos.Factura.objects.all().filter(estadofactura = 4, fechafactura__gte = self.request.POST.get('fechadesdeFacturas')).order_by('fechafactura')
                            
                            if self.request.POST.get('fechahastaFacturas'):
                                facturas_pagadas = modelos.Factura.objects.all().filter(estadofactura = 2, fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                                facturas_no_pagadas = modelos.Factura.objects.all().filter(estadofactura = 1, fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                                facturas_trans_bancaria = modelos.Factura.objects.all().filter(estadofactura = 3, fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                                factura_presencial = modelos.Factura.objects.all().filter(estadofactura = 4, fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                            
                            if self.request.POST.get('fechadesdeFacturas') and self.request.POST.get('fechahastaFacturas'):
                                facturas_pagadas = modelos.Factura.objects.all().filter(estadofactura = 2, fechafactura__gte = self.request.POST.get('fechadesdeFacturas'),fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                                facturas_no_pagadas = modelos.Factura.objects.all().filter(estadofactura = 1, fechafactura__gte = self.request.POST.get('fechadesdeFacturas'),fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                                facturas_trans_bancaria = modelos.Factura.objects.all().filter(estadofactura = 3, fechafactura__gte = self.request.POST.get('fechadesdeFacturas'),fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                                factura_presencial = modelos.Factura.objects.all().filter(estadofactura = 4, fechafactura__gte = self.request.POST.get('fechadesdeFacturas'),fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                    

                        if int(op) == 3:
                            for fp in facturas_pagadas:
                                ws.cell(row = cont, column = 2).value = fp.idfactura
                                ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 3).value = fp.cliente.nombre
                                ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 4).value = fp.giro
                                ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 5).value = fp.fechafactura
                                ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 6).value = fp.fechapago
                                ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 7).value = fp.estadofactura.descripcion
                                ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 8).value = fp.subtotal
                                ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 9).value = fp.iva
                                ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 10).value = fp.total
                                ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                cont += 1
                            
                        if int(op) == 4:
                            for fp in facturas_no_pagadas:
                                ws.cell(row = cont, column = 2).value = fp.idfactura
                                ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 3).value = fp.cliente.nombre
                                ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 4).value = fp.giro
                                ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 5).value = fp.fechafactura
                                ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 6).value = fp.fechapago
                                ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 7).value = fp.estadofactura.descripcion
                                ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 8).value = fp.subtotal
                                ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 9).value = fp.iva
                                ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 10).value = fp.total
                                ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                cont += 1

                        if int(op) == 5:
                            for fp in facturas_trans_bancaria:
                                ws.cell(row = cont, column = 2).value = fp.idfactura
                                ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 3).value = fp.cliente.nombre
                                ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 4).value = fp.giro
                                ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 5).value = fp.fechafactura
                                ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 6).value = fp.fechapago
                                ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 7).value = fp.estadofactura.descripcion
                                ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 8).value = fp.subtotal
                                ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 9).value = fp.iva
                                ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 10).value = fp.total
                                ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                cont += 1

                        if int(op) == 6:
                            for fp in factura_presencial:
                                ws.cell(row = cont, column = 2).value = fp.idfactura
                                ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 3).value = fp.cliente.nombre
                                ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 4).value = fp.giro
                                ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 5).value = fp.fechafactura
                                ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 6).value = fp.fechapago
                                ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 7).value = fp.estadofactura.descripcion
                                ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 8).value = fp.subtotal                            
                                ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 9).value = fp.iva
                                ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 10).value = fp.total
                                ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                cont += 1

                    if int(op) == 7 or int(op) == 8 or int(op) == 9 or int(op) == 10:
                        if rep_pedidos:
                            rep_pedidos = False
                            if bandera:
                                ws = wb.active
                                ws.title = "Informe Productos Solicitados"
                                bandera = False
                            else:
                                ws = wb.create_sheet("Informe Productos Solicitados")
                            
                            # TITULO DEL REPORTE
                            ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
                            ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws['B1'].fill = PatternFill(start_color = '3D3D3D', end_color = '3D3D3D', fill_type = "solid")
                            ws['B1'].font = Font(name = 'Calibri', size = 12, bold = True, color = 'FFFFFF')

                            ws['B1'] = "INFORME DE PRODUCTOS SOLICITADOS"
                            ws.row_dimensions[1].height = 45
                            ws.merge_cells('B1:K1')

                            # TAMAÑO COLUMNAS
                            ws.column_dimensions['B'].width = 15
                            ws.column_dimensions['C'].width = 22
                            ws.column_dimensions['D'].width = 22
                            ws.column_dimensions['E'].width = 22
                            ws.column_dimensions['F'].width = 22
                            ws.column_dimensions['G'].width = 22
                            ws.column_dimensions['H'].width = 22
                            ws.column_dimensions['I'].width = 22
                            ws.column_dimensions['J'].width = 15
                            ws.column_dimensions['K'].width = 15

                            # CAMPOS
                            ws['B3'] = "# PEDIDO"
                            ws['C3'] = "FECHA PEDIDO"
                            ws['D3'] = "ESTADO"
                            ws['E3'] = "FECHA ENTREGA"
                            ws['F3'] = "PRODUCTO"
                            ws['G3'] = "CANTIDAD"
                            ws['H3'] = "PROVEEDOR"
                            ws['I3'] = "SUBTOTAL"
                            ws['J3'] = "IVA"
                            ws['K3'] = "TOTAL"

                            # ESTILO CABEZERA
                            for i in range(2, 12):
                                ws.cell(row = 3, column = i).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = 3, column = i).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = 3, column = i).fill = PatternFill(start_color = 'A0A0A0', end_color = 'A0A0A0', fill_type = "solid")
                                ws.cell(row = 3, column = i).font = Font(name = 'Calibri', size = 12, bold = True)
                            
                            # DATOS
                            cont = 4

                        if tipousuario == 2:                  
                            pedidos_nuevos = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 1).order_by('fechapedido')
                            pedidos_en_transito = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 2).order_by('fechapedido')
                            pedidos_recibidos = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 3).order_by('fechapedido')
                            pedidos_rechazados = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 4).order_by('fechapedido')

                            if self.request.POST.get('fechadesdeSolicitudes'):
                                pedidos_nuevos = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 1, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes')).order_by('fechapedido')
                                pedidos_en_transito = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 2, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes')).order_by('fechapedido')
                                pedidos_recibidos = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 3, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes')).order_by('fechapedido')
                                pedidos_rechazados = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 4, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes')).order_by('fechapedido')
                            
                            if self.request.POST.get('fechahastaSolicitudes'):
                                pedidos_nuevos = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 1, fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                                pedidos_en_transito = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 2, fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                                pedidos_recibidos = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 3, fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                                pedidos_rechazados = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 4, fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                            
                            if self.request.POST.get('fechadesdeSolicitudes') and self.request.POST.get('fechahastaSolicitudes'):
                                pedidos_nuevos = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 1, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes'),fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                                pedidos_en_transito = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 2, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes'),fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                                pedidos_recibidos = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 3, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes'),fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                                pedidos_rechazados = modelos.Pedido.objects.all().filter(empleado = empleado.idempleado, estadopedido = 4, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes'),fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')

                        if tipousuario == 1:                  
                            pedidos_nuevos = modelos.Pedido.objects.all().filter(estadopedido = 1).order_by('fechapedido')
                            pedidos_en_transito = modelos.Pedido.objects.all().filter(estadopedido = 2).order_by('fechapedido')
                            pedidos_recibidos = modelos.Pedido.objects.all().filter(estadopedido = 3).order_by('fechapedido')
                            pedidos_rechazados = modelos.Pedido.objects.all().filter(estadopedido = 4).order_by('fechapedido')

                            if self.request.POST.get('fechadesdeSolicitudes'):
                                pedidos_nuevos = modelos.Pedido.objects.all().filter(estadopedido = 1, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes')).order_by('fechapedido')
                                pedidos_en_transito = modelos.Pedido.objects.all().filter(estadopedido = 2, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes')).order_by('fechapedido')
                                pedidos_recibidos = modelos.Pedido.objects.all().filter(estadopedido = 3, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes')).order_by('fechapedido')
                                pedidos_rechazados = modelos.Pedido.objects.all().filter(estadopedido = 4, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes')).order_by('fechapedido')
                            
                            if self.request.POST.get('fechahastaSolicitudes'):
                                pedidos_nuevos = modelos.Pedido.objects.all().filter(estadopedido = 1, fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                                pedidos_en_transito = modelos.Pedido.objects.all().filter(estadopedido = 2, fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                                pedidos_recibidos = modelos.Pedido.objects.all().filter(estadopedido = 3, fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                                pedidos_rechazados = modelos.Pedido.objects.all().filter(estadopedido = 4, fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                            
                            if self.request.POST.get('fechadesdeSolicitudes') and self.request.POST.get('fechahastaSolicitudes'):
                                pedidos_nuevos = modelos.Pedido.objects.all().filter(estadopedido = 1, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes'),fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                                pedidos_en_transito = modelos.Pedido.objects.all().filter(estadopedido = 2, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes'),fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                                pedidos_recibidos = modelos.Pedido.objects.all().filter(estadopedido = 3, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes'),fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                                pedidos_rechazados = modelos.Pedido.objects.all().filter(estadopedido = 4, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes'),fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')


                        if int(op) == 7:
                            for p in pedidos_nuevos:
                                detalles = modelos.Detallepedido.objects.all().filter(pedido = p.idpedido)
                                for d in detalles:
                                    ws.cell(row = cont, column = 2).value = p.idpedido
                                    ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 3).value = p.fechapedido
                                    ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 4).value = p.estadopedido.descripcion
                                    ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 5).value = p.fechaentrega
                                    ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 6).value = d.producto.descripcion
                                    ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 7).value = d.cantidad
                                    ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 8).value = p.proveedor.rubro
                                    ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 9).value = d.total
                                    ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 10).value = round(d.total*0.19)
                                    ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 11).value = round(d.total*0.19) + d.total
                                    ws.cell(row = cont, column = 11).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 11).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    cont += 1

                        if int(op) == 8:
                            for p in pedidos_en_transito:
                                detalles = modelos.Detallepedido.objects.all().filter(pedido = p.idpedido)
                                for d in detalles:
                                    ws.cell(row = cont, column = 2).value = p.idpedido
                                    ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 3).value = p.fechapedido
                                    ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 4).value = p.estadopedido.descripcion
                                    ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 5).value = p.fechaentrega
                                    ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 6).value = d.producto.descripcion
                                    ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 7).value = d.cantidad
                                    ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 8).value = p.proveedor.rubro
                                    ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 9).value = d.total
                                    ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 10).value = round(d.total*0.19)
                                    ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 11).value = round(d.total*0.19) + d.total
                                    ws.cell(row = cont, column = 11).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 11).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    cont += 1

                        if int(op) == 9:
                            for p in pedidos_recibidos:
                                detalles = modelos.Detallepedido.objects.all().filter(pedido = p.idpedido)
                                for d in detalles:
                                    ws.cell(row = cont, column = 2).value = p.idpedido
                                    ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 3).value = p.fechapedido
                                    ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 4).value = p.estadopedido.descripcion
                                    ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 5).value = p.fechaentrega
                                    ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 6).value = d.producto.descripcion
                                    ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 7).value = d.cantidad
                                    ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 8).value = p.proveedor.rubro
                                    ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 9).value = d.total
                                    ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 10).value = round(d.total*0.19)
                                    ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 11).value = round(d.total*0.19) + d.total
                                    ws.cell(row = cont, column = 11).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 11).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    cont += 1

                        if int(op) == 10:
                            for p in pedidos_rechazados:
                                detalles = modelos.Detallepedido.objects.all().filter(pedido = p.idpedido)
                                for d in detalles:
                                    ws.cell(row = cont, column = 2).value = p.idpedido
                                    ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 3).value = p.fechapedido
                                    ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 4).value = p.estadopedido.descripcion
                                    ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 5).value = p.fechaentrega
                                    ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 6).value = d.producto.descripcion
                                    ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 7).value = d.cantidad
                                    ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 8).value = p.proveedor.rubro
                                    ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 9).value = d.total
                                    ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 10).value = round(d.total*0.19)
                                    ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 11).value = round(d.total*0.19) + d.total
                                    ws.cell(row = cont, column = 11).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 11).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    cont += 1
                    
                    if int(op) == 11:
                        if bandera:
                            ws = wb.active
                            ws.title = "Informe Productos Recibidos"
                            bandera = False
                        else:
                            ws = wb.create_sheet("Informe Productos Recibidos")
                        
                        # TITULO DEL REPORTE
                        ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
                        ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                        ws['B1'].fill = PatternFill(start_color = '3D3D3D', end_color = '3D3D3D', fill_type = "solid")
                        ws['B1'].font = Font(name = 'Calibri', size = 12, bold = True, color = 'FFFFFF')

                        ws['B1'] = "INFORME DE PRODUCTOS RECIBIDOS"
                        ws.row_dimensions[1].height = 45
                        ws.merge_cells('B1:F1')
                        # TAMAÑO COLUMNAS
                        ws.column_dimensions['B'].width = 15
                        ws.column_dimensions['C'].width = 25
                        ws.column_dimensions['D'].width = 22
                        ws.column_dimensions['E'].width = 40
                        ws.column_dimensions['F'].width = 22
                       

                        # CAMPOS
                        ws['B3'] = "#"
                        ws['C3'] = "CODIGO"
                        ws['D3'] = "FECHA RECEPCION"
                        ws['E3'] = "PRODUCTO"
                        ws['F3'] = "CANTIDAD"
                        

                        # ESTILO CABEZERA
                        for i in range(2, 7):
                            ws.cell(row = 3, column = i).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = 3, column = i).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = 3, column = i).fill = PatternFill(start_color = 'A0A0A0', end_color = 'A0A0A0', fill_type = "solid")
                            ws.cell(row = 3, column = i).font = Font(name = 'Calibri', size = 12, bold = True)
                            
                        # DATOS
                        cont = 4

                        if tipousuario == 2:

                            productos_recibidos = modelos.Recepcionproducto.objects.all().filter(empleado= empleado.idempleado).order_by('id')

                            if self.request.POST.get('fechadesdeSolicitudes'):
                                productos_recibidos = modelos.Recepcionproducto.objects.all().filter(empleado= empleado.idempleado, fecharecepcion__gte = self.request.POST.get('fechadesdeSolicitudes')).order_by('id')

                            if self.request.POST.get('fechahastaSolicitudes'):
                                productos_recibidos = modelos.Recepcionproducto.objects.all().filter(empleado= empleado.idempleado, fecharecepcion__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('id')

                            if self.request.POST.get('fechadesdeSolicitudes') and self.request.POST.get('fechahastaSolicitudes'):
                                productos_recibidos = modelos.Recepcionproducto.objects.all().filter(empleado= empleado.idempleado,fecharecepcion__gte = self.request.POST.get('fechadesdeSolicitudes'), fecharecepcion__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('id')

                        if tipousuario == 1:

                            productos_recibidos = modelos.Recepcionproducto.objects.all().order_by('id')

                            if self.request.POST.get('fechadesdeSolicitudes'):
                                productos_recibidos = modelos.Recepcionproducto.objects.all().filter(fecharecepcion__gte = self.request.POST.get('fechadesdeSolicitudes')).order_by('id')

                            if self.request.POST.get('fechahastaSolicitudes'):
                                productos_recibidos = modelos.Recepcionproducto.objects.all().filter(fecharecepcion__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('id')

                            if self.request.POST.get('fechadesdeSolicitudes') and self.request.POST.get('fechahastaSolicitudes'):
                                productos_recibidos = modelos.Recepcionproducto.objects.all().filter(fecharecepcion__gte = self.request.POST.get('fechadesdeSolicitudes'), fecharecepcion__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('id')


                        for p in productos_recibidos:
                            
                            ws.cell(row = cont, column = 2).value = p.id
                            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 3).value = p.codigo
                            ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 4).value = p.fecharecepcion
                            ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 5).value = p.detallepedido.producto.descripcion
                            ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 6).value = p.detallepedido.cantidad
                            ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            cont += 1

            if tipousuario == 3:
                cliente = modelos.Cliente.objects.get(usuario = self.request.user.idusuario)
                rep_res = True
                rep_fac = True
                for op in opciones:

                    if int(op) == 1 or int(op) == 2:
                        if rep_res:
                            rep_res = False
                            if bandera:
                                ws = wb.active
                                ws.title = "Informe Reservas"
                                bandera = False
                            else:
                                ws = wb.create_sheet("Informe Reservas")
                            
                            # TITULO DEL REPORTE
                            ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
                            ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws['B1'].fill = PatternFill(start_color = '3D3D3D', end_color = '3D3D3D', fill_type = "solid")
                            ws['B1'].font = Font(name = 'Calibri', size = 12, bold = True, color = 'FFFFFF')

                            ws['B1'] = "INFORME DE RESERVAS"
                            ws.row_dimensions[1].height = 45
                            ws.merge_cells('B1:J1')

                            # TAMAÑO COLUMNAS
                            ws.column_dimensions['B'].width = 15
                            ws.column_dimensions['C'].width = 22
                            ws.column_dimensions['D'].width = 22
                            ws.column_dimensions['E'].width = 22
                            ws.column_dimensions['F'].width = 22
                            ws.column_dimensions['G'].width = 22
                            ws.column_dimensions['H'].width = 15
                            ws.column_dimensions['I'].width = 15
                            ws.column_dimensions['J'].width = 15

                            # CAMPOS
                            ws['B3'] = "RUT"
                            ws['C3'] = "HUESPED"
                            ws['D3'] = "FECHA DESDE"
                            ws['E3'] = "FECHA HASTA"
                            ws['F3'] = "DIAS"
                            ws['G3'] = "HABITACION"
                            ws['H3'] = "SUBTOTAL"
                            ws['I3'] = "IVA"
                            ws['J3'] = "TOTAL"

                            # ESTILO CABEZERA
                            for i in range(2, 11):
                                ws.cell(row = 3, column = i).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = 3, column = i).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = 3, column = i).fill = PatternFill(start_color = 'A0A0A0', end_color = 'A0A0A0', fill_type = "solid")
                                ws.cell(row = 3, column = i).font = Font(name = 'Calibri', size = 12, bold = True)
                            
                            # DATOS
                            cont = 4

                        reservas_activas = modelos.Huesped.objects.all().filter(cliente = cliente.idcliente, fechahasta__gte = date.today())
                        reservas_finalizadas = modelos.Huesped.objects.all().filter(cliente = cliente.idcliente, fechahasta__lte = date.today())
                        
                        if self.request.POST.get('fechadesdeReserva'):
                            reservas_activas = modelos.Huesped.objects.all().filter(cliente = cliente.idcliente, fechahasta__gte = date.today(), fechadesde__gte = self.request.POST.get('fechadesdeReserva'))
                            reservas_finalizadas = modelos.Huesped.objects.all().filter(cliente = cliente.idcliente, fechahasta__lte = date.today(), fechadesde__gte = self.request.POST.get('fechadesdeReserva'))

                        if self.request.POST.get('fechahastaReserva'):
                            reservas_activas = modelos.Huesped.objects.all().filter(cliente = cliente.idcliente, fechahasta__gte = date.today(), fechadesde__lte = self.request.POST.get('fechahastaReserva'))
                            reservas_finalizadas = modelos.Huesped.objects.all().filter(cliente = cliente.idcliente, fechahasta__lte = date.today(), fechadesde__lte = self.request.POST.get('fechahastaReserva'))

                        if self.request.POST.get('fechadesdeReserva') and self.request.POST.get('fechahastaReserva'):
                            reservas_activas = modelos.Huesped.objects.all().filter(cliente = cliente.idcliente, fechahasta__gte = date.today(), fechadesde__gte = self.request.POST.get('fechadesdeReserva'), fechadesde__lte = self.request.POST.get('fechahastaReserva'))
                            reservas_finalizadas = modelos.Huesped.objects.all().filter(cliente = cliente.idcliente, fechahasta__lte = date.today(), fechadesde__gte = self.request.POST.get('fechadesdeReserva'), fechadesde__lte = self.request.POST.get('fechahastaReserva'))
                        
                        if int(op) == 1:
                            for r in reservas_activas:
                                detalles = modelos.DetalleReserva.objects.all().filter(huesped = r.idhuesped)
                                for d in detalles:
                                    ws.cell(row = cont, column = 2).value = r.rut
                                    ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 3).value = r.nombre + ' ' + r.apellidopaterno + ' ' + r.apellidomaterno
                                    ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 4).value = r.fechadesde
                                    ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 5).value = r.fechahasta
                                    ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 6).value = d.dias
                                    ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 7).value = r.habitacion.numero
                                    ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 8).value = d.total
                                    ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 9).value = round(d.total*0.19)
                                    ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 10).value = round(d.total*0.19) + d.total
                                    ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    cont += 1
                        
                        if int(op) == 2:
                            for r in reservas_finalizadas:
                                detalles = modelos.DetalleReserva.objects.all().filter(huesped = r.idhuesped)
                                for d in detalles:
                                    ws.cell(row = cont, column = 2).value = r.rut
                                    ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 3).value = r.nombre + ' ' + r.apellidopaterno + ' ' + r.apellidomaterno
                                    ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 4).value = r.fechadesde
                                    ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 5).value = r.fechahasta
                                    ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 6).value = d.dias
                                    ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 7).value = r.habitacion.numero
                                    ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 8).value = d.total
                                    ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 9).value = round(d.total*0.19)
                                    ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 10).value = round(d.total*0.19) + d.total
                                    ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    cont += 1
                        
                    if int(op) == 3:
                
                        factura = modelos.Factura.objects.all().values_list('idfactura', flat=True).filter(cliente = cliente.idcliente, estadofactura = 2).order_by('fechapago')
                    

                        if self.request.POST.get('fechadesdeReserva'):
                            factura = modelos.Factura.objects.all().values_list('idfactura', flat=True).filter(cliente = cliente.idcliente, estadofactura = 2, fechapago__gte = self.request.POST.get('fechadesdeReserva')).order_by('fechapago')
                            

                        if self.request.POST.get('fechahastaReserva'):
                            factura = modelos.Factura.objects.all().values_list('idfactura', flat=True).filter(cliente = cliente.idcliente, estadofactura = 2, fechapago__lte = self.request.POST.get('fechahastaReserva')).order_by('fechapago')
                            
                        
                        if self.request.POST.get('fechadesdeReserva') and self.request.POST.get('fechahastaReserva'):
                            factura = modelos.Factura.objects.all().values_list('idfactura', flat=True).filter(cliente = cliente.idcliente, estadofactura = 2, fechapago__gte = self.request.POST.get('fechadesdeReserva'), fechapago__lte = self.request.POST.get('fechahastaReserva')).order_by('fechapago')
                            
                        compras_realizadas = modelos.Detallefactura.objects.all().filter(factura__in = factura, solicitudcompra__isnull = False).order_by('factura')


                        if bandera:
                            ws = wb.active
                            ws.title = "Informe Compras Realizadas"
                            bandera = False
                        else:
                            ws = wb.create_sheet("Informe Compras Realizadas")

                        # TITULO DEL REPORTE
                        ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
                        ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                        ws['B1'].fill = PatternFill(start_color = '3D3D3D', end_color = '3D3D3D', fill_type = "solid")
                        ws['B1'].font = Font(name = 'Calibri', size = 12, bold = True, color = 'FFFFFF')

                        ws['B1'] = "INFORME DE COMPRAS REALIZADAS"
                        ws.row_dimensions[1].height = 45
                        ws.merge_cells('B1:K1')

                        # TAMAÑO COLUMNA
                        ws.column_dimensions['B'].width = 15
                        ws.column_dimensions['C'].width = 22
                        ws.column_dimensions['D'].width = 22
                        ws.column_dimensions['E'].width = 22
                        ws.column_dimensions['F'].width = 22
                        ws.column_dimensions['G'].width = 40
                        ws.column_dimensions['H'].width = 22
                        ws.column_dimensions['I'].width = 15
                        ws.column_dimensions['J'].width = 15
                        ws.column_dimensions['K'].width = 15

                        # CAMPOS
                        ws['B3'] = "RUT"
                        ws['C3'] = "NOMBRE"
                        ws['D3'] = "APELLIDO PATERNO"
                        ws['E3'] = "APELLIDO MATERNO"
                        ws['F3'] = "FECHA COMPRA"
                        ws['G3'] = "SERVICIO"
                        ws['H3'] = "CANTIDAD"
                        ws['I3'] = "SUBTOTAL"
                        ws['J3'] = "IVA"
                        ws['K3'] = "TOTAL"

                        # ESTILO CABEZERA 
                        for i in range(2, 12):
                            ws.cell(row = 3, column = i).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = 3, column = i).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = 3, column = i).fill = PatternFill(start_color = 'A0A0A0', end_color = 'A0A0A0', fill_type = "solid")
                            ws.cell(row = 3, column = i).font = Font(name = 'Calibri', size = 12, bold = True)

                        
                        # DATOS
                        cont = 4

                        for cr in compras_realizadas:
                            ws.cell(row = cont, column = 2).value = cr.huesped.rut
                            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 3).value = cr.huesped.nombre
                            ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 4).value = cr.huesped.apellidopaterno
                            ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 5).value = cr.huesped.apellidomaterno
                            ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 6).value = cr.factura.fechapago
                            ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 7).value = cr.solicitudcompra.serviciocomedor.plato
                            ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 8).value = cr.solicitudcompra.cantidad
                            ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 9).value = cr.total
                            ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 10).value = round(cr.total*0.19)
                            ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws.cell(row = cont, column = 11).value = cr.total + round(cr.total*0.19)
                            ws.cell(row = cont, column = 11).alignment = Alignment(horizontal = "center", vertical = "center")
                            ws.cell(row = cont, column = 11).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            cont += 1

                    if int(op) == 4 or int(op) == 5 or int(op) == 6 or int(op) == 7:
                        if rep_fac:
                            rep_fac = False
                            if bandera:
                                ws = wb.active
                                ws.title = "Informe Facturas"
                                bandera = False
                            else:
                                ws = wb.create_sheet("Informe Facturas")
                            
                            # TITULO DEL REPORTE
                            ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
                            ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws['B1'].fill = PatternFill(start_color = '3D3D3D', end_color = '3D3D3D', fill_type = "solid")
                            ws['B1'].font = Font(name = 'Calibri', size = 12, bold = True, color = 'FFFFFF')

                            ws['B1'] = "INFORME DE FACTURAS"
                            ws.row_dimensions[1].height = 45
                            ws.merge_cells('B1:J1')

                            # TAMAÑO COLUMNAS
                            ws.column_dimensions['B'].width = 15
                            ws.column_dimensions['C'].width = 22
                            ws.column_dimensions['D'].width = 22
                            ws.column_dimensions['E'].width = 22
                            ws.column_dimensions['F'].width = 22
                            ws.column_dimensions['G'].width = 22
                            ws.column_dimensions['H'].width = 15
                            ws.column_dimensions['I'].width = 15
                            ws.column_dimensions['J'].width = 15

                            # CAMPOS
                            ws['B3'] = "#"
                            ws['C3'] = "CLIENTE"
                            ws['D3'] = "GIRO"
                            ws['E3'] = "FECHA EMITIDA"
                            ws['F3'] = "FECHA PAGO"
                            ws['G3'] = "ESTADO"
                            ws['H3'] = "SUBTOTAL"
                            ws['I3'] = "IVA"
                            ws['J3'] = "TOTAL"

                            # ESTILO CABEZERA
                            for i in range(2, 11):
                                ws.cell(row = 3, column = i).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = 3, column = i).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = 3, column = i).fill = PatternFill(start_color = 'A0A0A0', end_color = 'A0A0A0', fill_type = "solid")
                                ws.cell(row = 3, column = i).font = Font(name = 'Calibri', size = 12, bold = True)
                            
                            # DATOS
                            cont = 4

                        facturas_pagadas = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 2).order_by('fechafactura')
                        facturas_no_pagadas = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 1).order_by('fechafactura')
                        facturas_trans_bancaria = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 3).order_by('fechafactura')
                        factura_presencial = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 4).order_by('fechafactura')

                        if self.request.POST.get('fechadesdeFacturas'):
                            facturas_pagadas = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 2, fechafactura__gte = self.request.POST.get('fechadesdeFacturas')).order_by('fechafactura')
                            facturas_no_pagadas = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 1, fechafactura__gte = self.request.POST.get('fechadesdeFacturas')).order_by('fechafactura')
                            facturas_trans_bancaria = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 3, fechafactura__gte = self.request.POST.get('fechadesdeFacturas')).order_by('fechafactura')
                            factura_presencial = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 4, fechafactura__gte = self.request.POST.get('fechadesdeFacturas')).order_by('fechafactura')
                        
                        if self.request.POST.get('fechahastaFacturas'):
                            facturas_pagadas = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 2, fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                            facturas_no_pagadas = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 1, fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                            facturas_trans_bancaria = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 3, fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                            factura_presencial = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 4, fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                        
                        if self.request.POST.get('fechadesdeFacturas') and self.request.POST.get('fechahastaFacturas'):
                            facturas_pagadas = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 2, fechafactura__gte = self.request.POST.get('fechadesdeFacturas'),fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                            facturas_no_pagadas = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 1, fechafactura__gte = self.request.POST.get('fechadesdeFacturas'),fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                            facturas_trans_bancaria = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 3, fechafactura__gte = self.request.POST.get('fechadesdeFacturas'),fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')
                            factura_presencial = modelos.Factura.objects.all().filter(cliente = cliente.idcliente, estadofactura = 4, fechafactura__gte = self.request.POST.get('fechadesdeFacturas'),fechafactura__lte = self.request.POST.get('fechahastaFacturas')).order_by('fechafactura')


                        if int(op) == 4:
                            for fp in facturas_pagadas:
                                ws.cell(row = cont, column = 2).value = fp.idfactura
                                ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 3).value = fp.cliente.nombre
                                ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 4).value = fp.giro
                                ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 5).value = fp.fechafactura
                                ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 6).value = fp.fechapago
                                ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 7).value = fp.estadofactura.descripcion
                                ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 8).value = fp.subtotal
                                ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 9).value = fp.iva
                                ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 10).value = fp.total
                                ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                cont += 1
                            
                        if int(op) == 5:
                            for fp in facturas_no_pagadas:
                                ws.cell(row = cont, column = 2).value = fp.idfactura
                                ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 3).value = fp.cliente.nombre
                                ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 4).value = fp.giro
                                ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 5).value = fp.fechafactura
                                ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 6).value = fp.fechapago
                                ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 7).value = fp.estadofactura.descripcion
                                ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 8).value = fp.subtotal
                                ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 9).value = fp.iva
                                ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 10).value = fp.total
                                ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                cont += 1

                        if int(op) == 6:
                            for fp in facturas_trans_bancaria:
                                ws.cell(row = cont, column = 2).value = fp.idfactura
                                ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 3).value = fp.cliente.nombre
                                ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 4).value = fp.giro
                                ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 5).value = fp.fechafactura
                                ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 6).value = fp.fechapago
                                ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 7).value = fp.estadofactura.descripcion
                                ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 8).value = fp.subtotal
                                ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 9).value = fp.iva
                                ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 10).value = fp.total
                                ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                cont += 1

                        if int(op) == 7:
                            for fp in factura_presencial:
                                ws.cell(row = cont, column = 2).value = fp.idfactura
                                ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 3).value = fp.cliente.nombre
                                ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 4).value = fp.giro
                                ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 5).value = fp.fechafactura
                                ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 6).value = fp.fechapago
                                ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 7).value = fp.estadofactura.descripcion
                                ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 8).value = fp.subtotal                            
                                ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 9).value = fp.iva
                                ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = cont, column = 10).value = fp.total
                                ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                cont += 1

            if tipousuario == 4:

                proveedor = modelos.Proveedor.objects.get(usuario = self.request.user.idusuario)
                rep_sol_prov = True
                for op in opciones:

                    if int(op) == 1 or int(op) == 2 or int(op) == 3 or int(op) == 4:
                        if rep_sol_prov:
                            rep_sol_prov = False
                            if bandera:
                                ws = wb.active
                                ws.title = "Informe Productos Solicitados"
                                bandera = False
                            else:
                                ws = wb.create_sheet("Informe Productos Solicitados")
                            
                            # TITULO DEL REPORTE
                            ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
                            ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                            ws['B1'].fill = PatternFill(start_color = '3D3D3D', end_color = '3D3D3D', fill_type = "solid")
                            ws['B1'].font = Font(name = 'Calibri', size = 12, bold = True, color = 'FFFFFF')

                            ws['B1'] = "INFORME DE PRODUCTOS SOLICITADOS"
                            ws.row_dimensions[1].height = 45
                            ws.merge_cells('B1:J1')

                            # TAMAÑO COLUMNAS
                            ws.column_dimensions['B'].width = 15
                            ws.column_dimensions['C'].width = 22
                            ws.column_dimensions['D'].width = 22
                            ws.column_dimensions['E'].width = 22
                            ws.column_dimensions['F'].width = 22
                            ws.column_dimensions['G'].width = 22
                            ws.column_dimensions['H'].width = 22
                            ws.column_dimensions['I'].width = 15
                            ws.column_dimensions['J'].width = 15

                            # CAMPOS
                            ws['B3'] = "# PEDIDO"
                            ws['C3'] = "FECHA PEDIDO"
                            ws['D3'] = "ESTADO"
                            ws['E3'] = "FECHA ENTREGA"
                            ws['F3'] = "PRODUCTO"
                            ws['G3'] = "CANTIDAD"
                            ws['H3'] = "SUBTOTAL"
                            ws['I3'] = "IVA"
                            ws['J3'] = "TOTAL"

                            # ESTILO CABEZERA
                            for i in range(2, 11):
                                ws.cell(row = 3, column = i).alignment = Alignment(horizontal = "center", vertical = "center")
                                ws.cell(row = 3, column = i).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                ws.cell(row = 3, column = i).fill = PatternFill(start_color = 'A0A0A0', end_color = 'A0A0A0', fill_type = "solid")
                                ws.cell(row = 3, column = i).font = Font(name = 'Calibri', size = 12, bold = True)
                            
                            # DATOS
                            cont = 4
                        
                        pedidos_nuevos = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 1).order_by('fechapedido')
                        pedidos_en_transito = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 2).order_by('fechapedido')
                        pedidos_recibidos = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 3).order_by('fechapedido')
                        pedidos_rechazados = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 4).order_by('fechapedido')

                        if self.request.POST.get('fechadesdeSolicitudes'):
                            pedidos_nuevos = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 1, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes')).order_by('fechapedido')
                            pedidos_en_transito = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 2, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes')).order_by('fechapedido')
                            pedidos_recibidos = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 3, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes')).order_by('fechapedido')
                            pedidos_rechazados = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 4, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes')).order_by('fechapedido')
                        
                        if self.request.POST.get('fechahastaSolicitudes'):
                            pedidos_nuevos = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 1, fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                            pedidos_en_transito = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 2, fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                            pedidos_recibidos = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 3, fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                            pedidos_rechazados = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 4, fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                        
                        if self.request.POST.get('fechadesdeSolicitudes') and self.request.POST.get('fechahastaSolicitudes'):
                            pedidos_nuevos = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 1, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes'),fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                            pedidos_en_transito = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 2, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes'),fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                            pedidos_recibidos = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 3, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes'),fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')
                            pedidos_rechazados = modelos.Pedido.objects.all().filter(proveedor = proveedor.idproveedor, estadopedido = 4, fechapedido__gte = self.request.POST.get('fechadesdeSolicitudes'),fechapedido__lte = self.request.POST.get('fechahastaSolicitudes')).order_by('fechapedido')


                        if int(op) == 1:
                            for p in pedidos_nuevos:
                                detalles = modelos.Detallepedido.objects.all().filter(pedido = p.idpedido)
                                for d in detalles:
                                    ws.cell(row = cont, column = 2).value = p.idpedido
                                    ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 3).value = p.fechapedido
                                    ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 4).value = p.estadopedido.descripcion
                                    ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 5).value = p.fechaentrega
                                    ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 6).value = d.producto.descripcion
                                    ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 7).value = d.cantidad
                                    ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 8).value = d.total
                                    ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 9).value = round(d.total*0.19)
                                    ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 10).value = round(d.total*0.19) + d.total
                                    ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    cont += 1

                        if int(op) == 2:
                            for p in pedidos_en_transito:
                                detalles = modelos.Detallepedido.objects.all().filter(pedido = p.idpedido)
                                for d in detalles:
                                    ws.cell(row = cont, column = 2).value = p.idpedido
                                    ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 3).value = p.fechapedido
                                    ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 4).value = p.estadopedido.descripcion
                                    ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 5).value = p.fechaentrega
                                    ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 6).value = d.producto.descripcion
                                    ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 7).value = d.cantidad
                                    ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 8).value = d.total
                                    ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 9).value = round(d.total*0.19)
                                    ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 10).value = round(d.total*0.19) + d.total
                                    ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    cont += 1

                        if int(op) == 3:
                            for p in pedidos_recibidos:
                                detalles = modelos.Detallepedido.objects.all().filter(pedido = p.idpedido)
                                for d in detalles:
                                    ws.cell(row = cont, column = 2).value = p.idpedido
                                    ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 3).value = p.fechapedido
                                    ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    
                                    ws.cell(row = cont, column = 4).value = "Entregado"
                                    ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 5).value = p.fechaentrega
                                    ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 6).value = d.producto.descripcion
                                    ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 7).value = d.cantidad
                                    ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 8).value = d.total
                                    ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 9).value = round(d.total*0.19)
                                    ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 10).value = round(d.total*0.19) + d.total
                                    ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    cont += 1

                        if int(op) == 4:
                            for p in pedidos_rechazados:
                                detalles = modelos.Detallepedido.objects.all().filter(pedido = p.idpedido)
                                for d in detalles:
                                    ws.cell(row = cont, column = 2).value = p.idpedido
                                    ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 3).value = p.fechapedido
                                    ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 4).value = p.estadopedido.descripcion
                                    ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 5).value = p.fechaentrega
                                    ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 6).value = d.producto.descripcion
                                    ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 7).value = d.cantidad
                                    ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 8).value = d.total
                                    ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 9).value = round(d.total*0.19)
                                    ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    ws.cell(row = cont, column = 10).value = round(d.total*0.19) + d.total
                                    ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center", vertical = "center")
                                    ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
                                    cont += 1
                                                    
            nombre_archivo = "reporte-"+str(date.today())+".xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response
        else:
            messages.error(request, "Debe seleccionar al menos una opcion.")
            return redirect('informes')

       
# USUARIO

class ModificarPerfil(UserPassesTestMixin, SuccessMessageMixin, UpdateView):
    model = modelos.Usuario
    form_class = formularios.PerfilForm
    template_name = 'dashboard/usuario/perfil.html'
    success_message = 'Perfil Actualizado'
    context_object_name = 'usuario'

    def test_func(self):
        user = self.request.user
        if str(user.idusuario) == self.kwargs['pk']:
            return True

    def handle_no_permission(self):
        return redirect('dashboard')

    def get_success_url(self):

        return reverse_lazy('actualizar perfil', kwargs={'pk': self.request.user.idusuario})
    
    def get_context_data(self, **kwargs):
        usuario = self.request.user
        context = super(ModificarPerfil, self).get_context_data(**kwargs)

        if self.request.user.tipousuario.idtipousuario == 3:
            cliente = modelos.Cliente.objects.get(usuario = usuario.idusuario)
            context['cliente'] = cliente

        if self.request.user.tipousuario.idtipousuario == 4:
            proveedor = modelos.Proveedor.objects.get(usuario = usuario.idusuario)
            context['prov'] = proveedor

        return context

    def post(self, request, *args, **kwargs):

        if self.request.method == 'POST':
            if self.request.user.tipousuario.idtipousuario == 3:
                cliente = modelos.Cliente.objects.get(usuario = self.request.user.idusuario)
                cliente.rut = self.request.POST.get('rutempresa')
                cliente.nombre = self.request.POST.get('nombre_empresa')
                cliente.rubro = self.request.POST.get('rubro')
                cliente.direccion = self.request.POST.get('direccion_empresa')
                cliente.telefono = self.request.POST.get('telefono')
                cliente.save()

            if self.request.user.tipousuario.idtipousuario == 4:
                prov = modelos.Proveedor.objects.get(usuario = self.request.user.idusuario)
                prov.descripcion = self.request.POST.get('descripcion')
                prov.rubro = self.request.POST.get('rubro')
                prov.telefono = self.request.POST.get('telefono')
                prov.sitioweb = self.request.POST.get('sitio_web')
                prov.save()

        return super().post(request, *args, **kwargs)


# FORMULARIO CONTACTO

class Contacto(TemplateView):
    template_name = 'contacto/form.html'

    def post(self, request, *args, **kwargs):
        

        nombre = request.POST['nombre'] + " " + request.POST['apellido_paterno'] + " " + request.POST['apellido_materno']
        correo = request.POST['email']
        rut_empresa = request.POST['rut']
        nombre_empresa = request.POST['nombre_empresa']
        rubro = request.POST['rubro']
        direccion = request.POST['direccion']
        telefono = request.POST['telefono']
        mensaje = request.POST['mensaje']
        
        if request.method =='POST':
            recaptcha_token = request.POST.get('g-recaptcha-response')
            re_url = "https://www.google.com/recaptcha/api/siteverify"
            re_secret_key = '6LeyuaoZAAAAAEoQHh6p2W7UgmAyeO0GbpjmE4oW'
            re_data = {"secret" : re_secret_key, "response" : recaptcha_token}
            re_server_response = requests.post(url=re_url, data = re_data)
            re_json = json.loads(re_server_response.text)

           
            if re_json['success'] == False:
                messages.error(request, "Complete el captcha")
                ctx = {}
                ctx['nombre'] = request.POST['nombre']
                ctx['apellido_paterno'] = request.POST['apellido_paterno']
                ctx['apellido_materno'] = request.POST['apellido_materno']
                ctx['email'] = request.POST['email']
                ctx['rut'] = request.POST['rut']
                ctx['nombre_empresa'] = request.POST['nombre_empresa']
                ctx['rubro'] = request.POST['rubro']
                ctx['direccion'] = request.POST['direccion']
                ctx['telefono'] = request.POST['telefono']
                ctx['mensaje'] = request.POST['mensaje']
                return render(request, 'contacto/form.html', ctx)


            html = render_to_string('contacto/email.html', { 
                'nombre': nombre,
                'correo' : correo,
                'rut_empresa' : rut_empresa,
                'nombre_empresa' : nombre_empresa,
                'rubro' : rubro,
                'direccion' : direccion,
                'telefono' : telefono,
                'mensaje' : mensaje
                 })
            
            email = EmailMultiAlternatives(
                subject ='Formulario de Contacto',
                body  = html,
                to = [settings.EMAIL_HOST_USER],
                )
            email.content_subtype = "html"
            
            email.send()

            

        messages.success(request, "Mensaje enviado, la hostal se pondra en contacto con usted! =)")
        return HttpResponseRedirect('contacto')


